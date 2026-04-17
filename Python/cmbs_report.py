"""
CMBS INVESTOR REPORTING TOOL - Python Edition
==============================================
Replaces Excel VBA macro. No Excel COM required.
- Reads IRP (PIRPXLPU) and PIRPXLLR with openpyxl / xlrd
- Updates .xls CREFC files with xlrd + xlutils.copy
- Writes Excel run log (openpyxl)
- Run via CMBS_Reporting_Tool.bat or: python cmbs_report.py

Version history:
  1.0  Initial Python port from VBA v12.6
  1.1  Fix Gap 1: Property updates last block only (not all rows)
       Fix Gap 2: Supplemental A1 replaces date in text, not overwrites cell
       Fix Gap 3: Comp Finan uses col B as key + last-block detection
       Fix Gap 4: Res LOC handles continuation rows + 15 cols (A:O)
       Fix Gap 5: Financial A1 uses replace_date_in_text, Windows-safe format
  2.2  Fix Gap 6: Total Loan tab now also updates col L (col 12, 0-based 11)
       = Current Ending Scheduled Balance (IRP col 7). Was missing from v2.1.
       Restored filter_zero_balance() — removes paid-off deals (beg balance = 0).
  2.3  Loan ID normalization overhaul:
       + normalize_loan_id(): handles .0 decimals, NBSP, commas, sci notation
       + loan_id_variants(): multi-key matching for format mismatches
       + get_loan_ids() now indexes all variants into the lookup dict
       + All 8 matching sites (Periodic/Property/TotalLoan/CompFinan × xls/xlsx)
         now iterate loan_id_variants() instead of exact string match
       + WARN-level debug logging for first 5 unmatched rows per file
  2.4  Add missing irp_date_to_yyyymmdd() function — was crashing all
       Supplemental files at Total Loan tab processing.
"""

import os
import sys
import re
import shutil
import glob
import argparse
import traceback
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy
try:
    import win32com.client as win32
except ImportError:  # pragma: no cover - optional dependency
    win32 = None

import cmbs_config as cfg

# ── Globals ──────────────────────────────────────────────────────────────────
log_rows  = []   # List of dicts for the run log
file_rows = []   # List of dicts for per-file detail
run_start = None

# ── Logging helpers ──────────────────────────────────────────────────────────

def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"{ts}  {msg}")
    log_rows.append({"time": ts, "level": level, "msg": msg})

def log_err(msg):
    log(f"*** ERROR: {msg}", level="ERROR")

def _highlight_cell(cell):
    """In VALIDATION_MODE, highlight a cell yellow to mark it as script-touched."""
    if cfg.VALIDATION_MODE:
        cell.fill = YELLOW_FILL

def add_file_row(deal, det_date, servicer, file_type, status, path, note="",
                  matched=0, unmatched=0, total_rows=0):
    file_rows.append({
        "deal": deal, "det_date": det_date, "servicer": servicer,
        "file_type": file_type, "status": status, "path": path, "note": note,
        "matched": matched, "unmatched": unmatched, "total_rows": total_rows,
    })

def validate_output_file(dest, file_type):
    """Basic post-write validation. Returns list of warning strings."""
    warnings = []
    if not os.path.isfile(dest):
        warnings.append(f"{file_type}: Output file does not exist after save")
        return warnings
    size = os.path.getsize(dest)
    if size == 0:
        warnings.append(f"{file_type}: Output file is 0 bytes")
    elif size < 1024:
        warnings.append(f"{file_type}: Output file suspiciously small ({size} bytes)")
    return warnings


# ── Path helpers ─────────────────────────────────────────────────────────────

def prod_path(p):
    """Return production equivalent of a path (strips TEST_ROOT prefix)."""
    if cfg.TEST_MODE and p.startswith(cfg.TEST_ROOT):
        return cfg.PROD_ROOT + p[len(cfg.TEST_ROOT):]
    return p

def test_path(p):
    """Return test equivalent of a production path."""
    if cfg.TEST_MODE and p.startswith(cfg.PROD_ROOT) and not p.startswith(cfg.TEST_ROOT):
        return cfg.TEST_ROOT + p[len(cfg.PROD_ROOT):]
    return p

def servicer_to_folder(code):
    """K -> KeyBank, M -> Midland, TM/WF -> Trimont, else warn and pass-through."""
    result = cfg.SERVICER_FOLDER_MAP.get(code.upper(), None)
    if result is None:
        log(f"  WARNING: Unknown servicer code '{code}' — not in SERVICER_FOLDER_MAP. "
            f"Add to cmbs_config.py or use DEAL_FOLDER_OVERRIDES in deal_overrides.json.", "WARN")
        return code
    return result

def clean_filename(trans_id):
    """Remove characters illegal in Windows filenames."""
    return re.sub(r'[\\/:*?"<>|]', '', trans_id).strip()

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

# ── Date helpers ──────────────────────────────────────────────────────────────

def parse_det_date(det_date_str):
    """Parse '20260311' -> datetime. Returns None on failure."""
    try:
        return datetime.strptime(str(det_date_str)[:8], "%Y%m%d")
    except Exception:
        return None

def replace_date_in_text(txt, new_date_str):
    """
    FIX GAP 2 & 5: Mirrors VBA ReplaceDateInText exactly.
    Finds 'As of MM/DD/YY' in txt and replaces just the date portion.
    Falls back to scanning for any MM/DD/YYYY pattern.
    Returns the updated string (unchanged if no date found).
    """
    if not txt:
        return txt

    # Strategy 1: Look for 'As of ' followed by a date
    as_of_pos = txt.lower().find("as of ")
    if as_of_pos >= 0:
        date_start = as_of_pos + 6
        # Skip leading spaces
        while date_start < len(txt) and txt[date_start] == " ":
            date_start += 1
        # Find end of date (space, '(', newline)
        date_end = date_start
        while date_end < len(txt) and txt[date_end] not in (" ", "(", "\n", "\r"):
            date_end += 1
        return txt[:date_start] + new_date_str + txt[date_end:]

    # Strategy 2: Find any MM/DD/YYYY or M/D/YY pattern
    m = re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', txt)
    if m:
        return txt[:m.start()] + new_date_str + txt[m.end():]

    return txt

def format_date_short(dt):
    """Return mm/dd/yy string matching VBA Format(dt, 'mm/dd/yy'). Windows-safe."""
    return dt.strftime("%m/%d/%y")

def format_date_long(dt):
    """Return mm/dd/yyyy string matching VBA Format(dt, 'mm/dd/yyyy'). Windows-safe."""
    return dt.strftime("%m/%d/%Y")

def irp_date_to_yyyymmdd(value):
    """
    Convert an IRP date value to YYYYMMDD numeric string for Total Loan col O.
    Handles datetime objects, date objects, and string dates.
    Returns the value as-is if it cannot be parsed (defensive).
    """
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")
    s = str(value).replace("-", "").replace("/", "").strip()
    if len(s) >= 8 and s[:8].isdigit():
        return s[:8]
    return value

# ── IRP reading ───────────────────────────────────────────────────────────────

def read_irp(irp_path):
    """
    Read GANADOWN PIRPXLPU sheet.
    Returns list-of-lists. Index 0 of each row = None (padding so col 1 = index 1).
    data[0] = header row, data[1..] = data rows.
    """
    log(f"  Opening IRP: {os.path.basename(irp_path)}")
    wb = openpyxl.load_workbook(irp_path, read_only=True, data_only=True)
    ws = wb["PIRPXLPU"]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([None] + list(row))
    wb.close()
    log(f"  IRP: {len(data)-1} data rows, {len(data[0])-1} columns")
    return data

# Expected header fragments at key IRP column positions (case-insensitive substring match)
_IRP_HEADER_CHECKS = {
    cfg.IRP_COL_TRANS_ID:   "Transaction ID",
    cfg.IRP_COL_LOAN_ID:    "Loan ID",
    cfg.IRP_COL_END_BAL:    "Ending Scheduled",
    cfg.IRP_COL_DIST_DATE:  "Distribution Date",
    cfg.IRP_COL_BEG_BAL:    "Beginning Scheduled",
    cfg.IRP_COL_PAID_THRU:  "Paid Through",
    cfg.IRP_COL_SERVICER:   "Servicer",
}

def validate_irp_columns(irp_data):
    """Warn if IRP header labels don't match expected column positions."""
    if not irp_data:
        return
    header = irp_data[0]
    warnings = []
    for col_idx, expected_fragment in _IRP_HEADER_CHECKS.items():
        if col_idx >= len(header):
            warnings.append(f"    Col {col_idx}: MISSING (header has only {len(header)-1} columns)")
            continue
        actual = str(header[col_idx] or "").strip()
        if expected_fragment.lower() not in actual.lower():
            warnings.append(f"    Col {col_idx}: expected '{expected_fragment}' but found '{actual}'")
    if warnings:
        log("  WARNING: IRP column layout may have changed:", "WARN")
        for w in warnings:
            log(w, "WARN")
        log("  Review IRP_COL_* values in cmbs_config.py if data looks wrong.", "WARN")
    else:
        log("  IRP column validation passed.")

def validate_overrides():
    """Check that deal_overrides.json folder paths exist on disk."""
    overrides = getattr(cfg, "DEAL_FOLDER_OVERRIDES", {})
    if not overrides:
        return
    log("  Validating deal override paths...")
    for tid, path in overrides.items():
        if not os.path.isdir(path):
            log(f"    WARNING: Override path does not exist: {tid} -> {path}", "WARN")
        else:
            log(f"    OK: {tid}")

def read_pirpxllr(pirp_path):
    """
    Read PIRPXLLR LL_Res_LOC sheet.
    Returns list-of-lists (same None-padded convention).
    """
    wb = openpyxl.load_workbook(pirp_path, read_only=True, data_only=True)
    ws = wb[cfg.PIRPXLLR_SHEET]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([None] + list(row))
    wb.close()
    log(f"  PIRPXLLR: {len(data)-1} rows on {cfg.PIRPXLLR_SHEET}")
    return data


def build_prospectus_id_map(pirp_data, trans_id):
    """
    Build {normalized_loan_id: prospectus_loan_id} for a given deal
    from PIRPXLLR (LL_Res_LOC) using Transaction ID (col A),
    Loan ID (col C), and Prospectus Loan ID (col D).
    """
    mapping = {}
    if not pirp_data:
        return mapping

    for row in pirp_data[1:]:
        tid = str(row[cfg.PIRPXLLR_TRANS_COL] or "").strip()
        if tid != trans_id:
            continue
        loan_norm = normalize_loan_id(row[cfg.PIRPXLLR_LOAN_COL])
        prosp_id = row[cfg.PIRPXLLR_PROSP_COL] if len(row) > cfg.PIRPXLLR_PROSP_COL else None
        if loan_norm and prosp_id not in (None, ""):
            mapping.setdefault(loan_norm, prosp_id)

    return mapping

# ── Deal discovery ────────────────────────────────────────────────────────────

def build_trans_map(irp_data):
    """
    Returns {trans_id: [irp_row_index, ...]} from PIRPXLPU.
    Skips rows with no Loan ID (summary/total rows).
    """
    trans_map = {}
    for i, row in enumerate(irp_data[1:], start=1):
        tid  = str(row[cfg.IRP_COL_TRANS_ID] or "").strip()
        loan = str(row[cfg.IRP_COL_LOAN_ID]  or "").strip()
        if not tid or not loan:
            continue
        trans_map.setdefault(tid, []).append(i)
    return trans_map

def filter_zero_balance(trans_map, irp_data):
    """
    Remove deals where ALL rows have Beginning Scheduled Balance = 0.
    These are paid-off / terminated deals. Restored in v2.2.
    Returns (active_map, removed_list).
    """
    active, removed = {}, []
    for tid, rows in trans_map.items():
        total = sum(float(irp_data[r][cfg.IRP_COL_BEG_BAL] or 0) for r in rows)
        if total == 0:
            log(f"  FILTERED (zero balance): {tid}")
            removed.append((tid, "zero Beginning Scheduled Balance"))
        else:
            active[tid] = rows
    log(f"  Zero-balance filter: {len(removed)} removed, {len(active)} remaining")
    return active, removed

def filter_by_pirpxllr(trans_map, pirp_data):
    """
    Remove deals with no Paid Through Date in PIRPXLLR col F.
    Deals not found in PIRPXLLR at all are kept (same as macro).
    Returns (active_map, removed_list).
    """
    ptd_map = {}
    for row in pirp_data[1:]:
        tid = str(row[cfg.PIRPXLLR_TRANS_COL] or "").strip()
        ptd = row[cfg.PIRPXLLR_PTD_COL]
        if tid:
            ptd_map.setdefault(tid, []).append(ptd)

    active, removed = {}, []
    for tid, rows in trans_map.items():
        if tid in ptd_map and all(v is None for v in ptd_map[tid]):
            log(f"  FILTERED (no Paid Through Date): {tid}")
            removed.append((tid, "no Paid Through Date"))
        else:
            active[tid] = rows

    log(f"  Filter: {len(removed)} removed, {len(active)} active")
    return active, removed

def get_det_date(irp_data, trans_id):
    """Return determination date string YYYYMMDD for a deal."""
    for row in irp_data[1:]:
        if str(row[cfg.IRP_COL_TRANS_ID] or "").strip() == trans_id:
            d = row[cfg.IRP_COL_DET_DATE]
            if d is None:
                return ""
            if isinstance(d, (datetime, date)):
                return d.strftime("%Y%m%d")
            return str(d).replace("-","").replace("/","")[:8]
    return ""

_tracking_list_cache = None

def sync_tracking_list():
    """
    Refresh the local Active Conduit Pool Tracking List from the master on
    S:\\Reporting. Writes a new dated xlsx alongside cmbs_config.py using the
    column map TRACKING_LIST_COL_MAP. Gracefully returns None if the master
    isn't reachable (falls back to most recent local copy).
    """
    global _tracking_list_cache
    source = getattr(cfg, "TRACKING_LIST_SOURCE", "")
    if not source or not os.path.isfile(source):
        log(f"  Tracking list master not accessible: {source or '(not configured)'}")
        log(f"    Will use most recent local copy instead")
        return None

    config_dir = os.path.dirname(os.path.abspath(cfg.__file__))
    today = datetime.now()
    date_str = f"{today.month}.{today.day}.{today.year}"
    dest_path = os.path.join(
        config_dir, f"Active Conduit Pool Tracking List {date_str}.xlsx"
    )

    try:
        src_wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
        src_ws = src_wb[cfg.TRACKING_LIST_SHEET]

        dest_wb = openpyxl.Workbook()
        dest_ws = dest_wb.active
        dest_ws.title = cfg.TRACKING_LIST_SHEET

        header_row = cfg.TRACKING_LIST_HEADER_ROW
        col_map = cfg.TRACKING_LIST_COL_MAP

        for src_col, dest_col in col_map:
            dest_ws.cell(header_row, dest_col).value = src_ws.cell(header_row, src_col).value

        data_rows = 0
        dest_row = header_row + 1
        for src_row in src_ws.iter_rows(min_row=header_row + 1, values_only=True):
            pool = src_row[cfg.TRACKING_LIST_POOL_COL - 1] if len(src_row) >= cfg.TRACKING_LIST_POOL_COL else None
            if not pool:
                continue
            for src_col, dest_col in col_map:
                val = src_row[src_col - 1] if len(src_row) >= src_col else None
                dest_ws.cell(dest_row, dest_col).value = val
            dest_row += 1
            data_rows += 1

        src_wb.close()
        dest_wb.save(dest_path)
        log(f"  Tracking list synced: {data_rows} pools from master -> {os.path.basename(dest_path)}")
        _tracking_list_cache = None  # force reload on next _load_tracking_list()
        return dest_path
    except Exception as e:
        log_err(f"  Tracking list sync failed: {e}")
        return None

def _load_tracking_list():
    """
    Load the Active Conduit Pool Tracking List into a {pool: servicer_code} dict.
    Tracking list lives alongside cmbs_config.py and contains columns:
    Master Servicer, Pool, Cashiered/Non-cashiered, GID, Loan Number, Borrower Name.
    Used as a fallback when IRP col EC (Master Servicer) is empty.
    Returns {} silently if the file is missing or unreadable.
    """
    global _tracking_list_cache
    if _tracking_list_cache is not None:
        return _tracking_list_cache

    _tracking_list_cache = {}
    config_dir = os.path.dirname(os.path.abspath(cfg.__file__))
    pattern = os.path.join(config_dir, cfg.TRACKING_LIST_GLOB)
    matches = glob.glob(pattern)
    if not matches:
        log(f"  Tracking list not found (glob: {cfg.TRACKING_LIST_GLOB}) — servicer fallback disabled")
        return _tracking_list_cache

    # Pick most recently modified file (filename dates like "4.17.2026" don't sort correctly)
    matches.sort(key=os.path.getmtime)
    path = matches[-1]
    log(f"  Loading tracking list: {os.path.basename(path)}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[cfg.TRACKING_LIST_SHEET]
        unknown_names = set()
        for row in ws.iter_rows(min_row=cfg.TRACKING_LIST_HEADER_ROW + 1, values_only=True):
            name = row[cfg.TRACKING_LIST_SERVICER_COL - 1]
            pool = row[cfg.TRACKING_LIST_POOL_COL - 1]
            if not pool:
                continue
            pool = str(pool).strip()
            name = str(name).strip() if name else ""
            code = cfg.TRACKING_SERVICER_TO_CODE.get(name, "")
            if not code and name:
                unknown_names.add(name)
                continue
            _tracking_list_cache.setdefault(pool, code)
        wb.close()
        log(f"  Tracking list: {len(_tracking_list_cache)} pools indexed")
        for n in sorted(unknown_names):
            log(f"  WARNING: Tracking list servicer name '{n}' not in TRACKING_SERVICER_TO_CODE")
    except Exception as e:
        log_err(f"  Could not read tracking list '{path}': {e}")
    return _tracking_list_cache

def get_servicer(irp_data, trans_id):
    """
    Return Master Servicer code for a deal.
    Falls back to the Active Conduit Pool Tracking List when the IRP's
    Master Servicer column is empty (common for newly added deals).
    """
    for row in irp_data[1:]:
        if str(row[cfg.IRP_COL_TRANS_ID] or "").strip() == trans_id:
            svc = str(row[cfg.IRP_COL_SERVICER] or "").strip()
            if svc:
                return svc
            break
    tracking = _load_tracking_list()
    code = tracking.get(trans_id, "")
    if code:
        log(f"     Servicer resolved via tracking list: {trans_id} -> {code}")
    return code

def normalize_loan_id(value):
    """
    Robust loan-id normalizer for IRP + CREFC matching.

    Accepts None, strings, ints, and floats safely.
    - Strips leading/trailing whitespace
    - Replaces NBSP with regular spaces
    - Removes surrounding quotes
    - Removes commas
    - Removes internal spaces
    - Converts numeric-looking values (incl. scientific notation) to
      integer strings when they are effectively whole numbers
    - Preserves alphanumeric content (letters + digits) when present
    """
    if value is None:
        return ""

    # Fast path for plain ints
    if isinstance(value, int):
        return str(value)

    s = str(value)

    # Normalize whitespace and quotes
    s = s.replace("\u00A0", " ")
    s = s.strip()
    s = s.strip("'").strip('"')

    if not s:
        return ""

    # Remove commas and collapse internal whitespace
    s = s.replace(",", "")
    s = re.sub(r"\s+", "", s)

    if not s:
        return ""

    # If it's numeric-looking (including scientific notation), try to
    # convert integer-like values to a clean integer string.
    numeric_pattern = r"^[+-]?\d+(\.\d+)?([Ee][+-]?\d+)?$"
    if re.fullmatch(numeric_pattern, s):
        try:
            d = Decimal(s)
            if d == d.to_integral_value():
                return str(int(d))
        except (InvalidOperation, ValueError):
            # Fall through to alphanumeric handling
            pass

    # For mixed / alphanumeric IDs, keep letters and digits, drop punctuation.
    alnum = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    if alnum:
        return alnum

    # Fallback: cleaned, upper-cased string
    return s.upper()

def loan_id_variants(value):
    """
    Return an ordered list of possible normalized lookup keys for a loan ID.

    Order and behaviour:
    - Normalized base value first (from normalize_loan_id)
    - Cleaned raw upper-case form second (no spaces/commas)
    - Leading-zero-stripped numeric variant last, and only when the
      normalized base is fully numeric
    """
    base = normalize_loan_id(value)

    if not base:
        return []

    variants = []

    def _add(v):
        if v and v not in variants:
            variants.append(v)

    # 1) Normalized base first
    _add(base)

    # 2) Cleaned upper-case version of the raw value (spaces/commas removed)
    raw = str(value).replace("\u00A0", " ").strip().strip("'").strip('"')
    raw = raw.replace(",", "")
    raw = re.sub(r"\s+", "", raw)
    raw_up = raw.upper()
    _add(raw_up)

    # 3) Numeric-only variants last, but ONLY when the normalized base is
    #    purely digits. This avoids creating numeric keys from mixed
    #    alphanumeric IDs like A12345 or L-00123A.
    if base.isdigit():
        stripped = base.lstrip("0") or "0"
        _add(stripped)

    return variants

def get_loan_ids(irp_data, trans_id):
    """
    Return dict {loan_id_variant: row_index} for a deal.
    Dual-key: tries col 3 (Loan ID) then col 154 (Loan Number).
    Indexes all useful variants from loan_id_variants() so that
    different file formats can still resolve to the same IRP row.
    """
    result = {}
    for i, row in enumerate(irp_data[1:], start=1):
        if str(row[cfg.IRP_COL_TRANS_ID] or "").strip() != trans_id:
            continue
        for col in (cfg.IRP_COL_LOAN_ID, cfg.IRP_COL_LOAN_NUM):
            raw = row[col]
            for lid in loan_id_variants(raw):
                if lid:
                    result.setdefault(lid, i)
    return result

# ── S: drive folder resolution ────────────────────────────────────────────────

def extract_series(trans_id):
    """'BANK 2019-BNK23' -> '2019-BNK23'"""
    m = re.search(r'(\d{4}[-_]\S+)', trans_id)
    return m.group(1) if m else trans_id

def find_deal_folder(base_path, series):
    """Scan base_path for any folder whose name contains series."""
    if not os.path.isdir(base_path):
        return ""
    for name in os.listdir(base_path):
        if series.lower() in name.lower():
            full = os.path.join(base_path, name)
            if os.path.isdir(full):
                return full + os.sep
    return ""

def resolve_output_folder(trans_id, servicer, deal_tracker_path=None):
    """
    Find the deal's S: drive production folder.
    1. DEAL_FOLDER_OVERRIDES dict in cmbs_config (highest priority)
    2. Deal Tracker xlsm lookup (if path provided)
    3. Scan S:/Lenders/{servicer}/Reporting/ by series
    Test redirect happens separately in build_crefc_folder.
    """
    # -- Option 1: Manual override dict --
    overrides = getattr(cfg, "DEAL_FOLDER_OVERRIDES", {})
    if trans_id in overrides:
        folder = overrides[trans_id] + os.sep
        log(f"     Folder override: {folder}")
        return folder

    # -- Option 2: Deal Tracker --
    if deal_tracker_path and os.path.isfile(deal_tracker_path):
        folder = _lookup_deal_tracker(deal_tracker_path, trans_id)
        if folder:
            return folder

    if not servicer:
        log(f"     Folder resolve FAILED: no servicer code in IRP for {trans_id}")
        log(f"       Fix: add servicer code to Deal Tracker, or add to DEAL_FOLDER_OVERRIDES in cmbs_config.py")
        return ""
    lender = servicer_to_folder(servicer)
    base   = os.path.join(cfg.PROD_ROOT, lender, "Reporting", "")
    series = extract_series(trans_id)
    log(f"     Folder resolve: servicer={servicer} -> lender={lender}, series={series}")
    log(f"       scanning: {base}")
    folder = find_deal_folder(base, series)
    if folder:
        log(f"       found: {folder}")
        return folder

    # No existing folder: auto-create new-deal folder as "<series> <ticker>"
    # e.g. "MSC 2021-L5" -> "2021-L5 MSC"
    ticker = trans_id.split()[0] if " " in trans_id else ""
    new_name = f"{series} {ticker}".strip() if ticker else trans_id
    constructed = os.path.join(base, new_name, "")
    log(f"     Folder NOT found by scan (series '{series}' not in any subfolder name)")
    if cfg.TEST_MODE:
        log(f"       TEST MODE — would create new deal folder: {constructed}")
        return constructed
    try:
        os.makedirs(constructed, exist_ok=True)
        log(f"       Created new deal folder: {constructed}")
    except Exception as e:
        log_err(f"       Could not create new deal folder '{constructed}': {e}")
    return constructed

def _lookup_deal_tracker(xlsm_path, trans_id):
    """Read Deal Tracker sheet from the .xlsm to find the deal folder."""
    try:
        wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
        if cfg.DEAL_TRACKER_SHEET not in wb.sheetnames:
            wb.close()
            return ""
        ws = wb[cfg.DEAL_TRACKER_SHEET]
        for row in ws.iter_rows(min_row=cfg.DT_FIRST_DATA_ROW, values_only=True):
            lender_code = str(row[cfg.DT_LENDER_COL - 1] or "").strip()
            dt_trans    = str(row[cfg.DT_TRANS_COL  - 1] or "").strip()
            desc        = str(row[cfg.DT_DESC_COL   - 1] or "").strip()
            if dt_trans.lower() == trans_id.lower() and lender_code:
                lender = servicer_to_folder(lender_code)
                series = extract_series(trans_id)
                ticker = trans_id.split()[0] if " " in trans_id else ""
                name   = f"{series} {ticker}"
                if desc:
                    name += f" ({desc})"
                folder = os.path.join(cfg.PROD_ROOT, lender, "Reporting", name, "")
                wb.close()
                return folder
        wb.close()
    except Exception as e:
        log(f"     Deal Tracker lookup failed: {e}")
    return ""

def find_crefc_subfolder(deal_folder):
    """Return the CREFC subfolder, trying multiple naming variants."""
    for variant in cfg.CREFC_FOLDER_VARIANTS:
        p = os.path.join(deal_folder, variant, "")
        if os.path.isdir(p):
            log(f"     CREFC subfolder found: {variant}")
            return p
    log(f"     CREFC subfolder: none of {cfg.CREFC_FOLDER_VARIANTS} found in {deal_folder} — defaulting to CREFC (will be created)")
    return os.path.join(deal_folder, "CREFC", "")

def build_crefc_folder(deal_folder, det_date_str):
    """
    Build + create the output CREFC/mm.yyyy folder.
    In TEST_MODE, rewrites S:/Lenders/... -> S:/Lenders/Z. CMBS Test/...
    """
    crefc        = find_crefc_subfolder(deal_folder)
    month_folder = _det_date_to_month_folder(det_date_str)
    out          = os.path.join(crefc, month_folder, "")
    if cfg.TEST_MODE:
        out = test_path(out)
    ensure_dir(out)
    return out

def _det_date_to_month_folder(det_date_str):
    """'20260311' -> '03.2026'"""
    try:
        d = datetime.strptime(str(det_date_str)[:8], "%Y%m%d")
        return d.strftime("%m.%Y")
    except Exception:
        return det_date_str[:6]

def find_prev_month_folder(crefc_parent, det_date_str):
    """
    Find the previous month's CREFC subfolder.
    In TEST_MODE, go straight to production — test folders never have
    prior-month source files.
    """
    try:
        d = datetime.strptime(str(det_date_str)[:8], "%Y%m%d")
    except Exception:
        return ""
    prev      = d - relativedelta(months=1)
    month_str = prev.strftime("%m")

    # Patterns to try: 02.2026, 02.26, 2.2026, 2.26
    patterns = [prev.strftime("%m.%Y"), prev.strftime("%m.%y")]
    m_stripped = str(int(month_str))
    patterns += [f"{m_stripped}.{prev.strftime('%Y')}", f"{m_stripped}.{prev.strftime('%y')}"]

    def _scan(parent):
        if not os.path.isdir(parent):
            log(f"     prev_month scan: folder does not exist: {parent}")
            return ""
        log(f"     prev_month scan: looking in {parent}")
        log(f"       patterns tried: {patterns}")
        for pat in patterns:
            p = os.path.join(parent, pat, "")
            if os.path.isdir(p):
                log(f"       FOUND: {pat}")
                return p
        # Loose scan: any folder starting with the 2-digit month number
        subdirs = [n for n in sorted(os.listdir(parent)) if os.path.isdir(os.path.join(parent, n))]
        log(f"       exact match failed. subfolders in parent: {subdirs}")
        for name in subdirs:
            if name.startswith(month_str):
                log(f"       loose match: {name}")
                return os.path.join(parent, name, "")
        log(f"       no previous month folder found in {parent}")
        return ""

    # In TEST_MODE, skip test path entirely — go straight to production.
    # Prior-month source files only exist on production S: drive.
    if cfg.TEST_MODE:
        prod_parent = prod_path(crefc_parent)
        log(f"     Prior-month: scanning production (TEST_MODE)")
        result = _scan(prod_parent)
        if result:
            return result
        # Try sibling CREFC variant folders on the production deal folder
        deal_folder = os.path.dirname(prod_parent.rstrip(os.sep))
        for variant in cfg.CREFC_FOLDER_VARIANTS:
            vp = os.path.join(deal_folder, variant, "")
            if vp != prod_parent:
                result = _scan(vp)
                if result:
                    log(f"     Prev month: production variant ({variant}) -> {result}")
                    return result
        return ""

    # Production mode: scan the passed-in path directly
    return _scan(crefc_parent)

# ── .xls file helpers (xlrd + xlutils.copy) ──────────────────────────────────

def xls_cell(rb_sheet, row_0, col_0):
    """Read a cell value from xlrd sheet (0-based row/col). Converts dates."""
    try:
        cell = rb_sheet.cell(row_0, col_0)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, rb_sheet.book.datemode)
        return cell.value
    except Exception:
        return None

def _filename_lookup_names(trans_id):
    """Return list of names to try for file lookups (override first, then clean trans_id)."""
    overrides = getattr(cfg, "DEAL_FILENAME_OVERRIDES", {})
    names = []
    if trans_id in overrides:
        names.append(overrides[trans_id])
    names.append(clean_filename(trans_id))
    return names

def find_file(folder, file_type, trans_id):
    """
    Find a CREFC Periodic / Property / Supplemental file in folder.
    Handles both naming conventions:
      - Old (VBA underscore):  CREFC_Periodic_MSC 2020-L4.xls
      - New (date-prefix):     2026.02 CREFC Periodic File MSC 2020-L4.xls
    file_type = "Periodic" | "Property" | "Supplemental"
    Uses DEAL_FILENAME_OVERRIDES when IRP trans_id doesn't match filenames (e.g. BANK5 20255YR15 vs 2025-5YR15).
    """
    prefix = f"CREFC_{file_type}_"
    log(f"       find_file [{file_type}]: scanning {folder}")

    for clean in _filename_lookup_names(trans_id):
        if clean != clean_filename(trans_id):
            log(f"         trying filename override: '{clean}'")

        # Pattern 1: old underscore style — prefer .xlsx over .xls
        for ext in (".xlsx", ".xls"):
            p = os.path.join(folder, f"{prefix}{clean}{ext}")
            log(f"         try P1: {os.path.basename(p)} -> {'FOUND' if os.path.isfile(p) else 'not found'}")
            if os.path.isfile(p):
                return p

        # Pattern 2: date-prefix style — prefer .xlsx when multiple matches
        for pat in (f"*CREFC {file_type} File*{clean}*", f"*CREFC {file_type}*{clean}*"):
            hits = sorted(glob.glob(os.path.join(folder, pat)))
            log(f"         try P2: '{pat}' -> {[os.path.basename(h) for h in hits] if hits else 'no match'}")
            if hits:
                xlsx_hits = [h for h in hits if h.lower().endswith(".xlsx")]
                return (xlsx_hits or hits)[-1]

        # Pattern 3: broad underscore glob — prefer .xlsx when multiple matches
        hits = sorted(glob.glob(os.path.join(folder, f"{prefix}*{clean}*")))
        log(f"         try P3: '{prefix}*{clean}*' -> {[os.path.basename(h) for h in hits] if hits else 'no match'}")
        if hits:
            xlsx_hits = [h for h in hits if h.lower().endswith(".xlsx")]
            return (xlsx_hits or hits)[-1]

        # Pattern 4: loose type+clean (same style as find_financial_file) — prefer .xlsx
        loose_pat = f"*{file_type}*{clean}*"
        hits = sorted(glob.glob(os.path.join(folder, loose_pat)))
        log(f"         try P4: '{loose_pat}' -> {[os.path.basename(h) for h in hits] if hits else 'no match'}")
        if hits:
            xlsx_hits = [h for h in hits if h.lower().endswith(".xlsx")]
            return (xlsx_hits or hits)[0]

    # List what IS in the folder so we can see the actual filenames
    try:
        all_files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
        log(f"         folder contents ({len(all_files)} files): {all_files[:10]}")
    except Exception as e:
        log(f"         could not list folder: {e}")
    return ""

def find_financial_file(folder, trans_id):
    """Find 'YYYY.MM CREFC Financial File XXXX.xls(x)' in folder. Uses DEAL_FILENAME_OVERRIDES when needed."""
    log(f"       find_financial_file: scanning {folder}")
    for clean in _filename_lookup_names(trans_id):
        if clean != clean_filename(trans_id):
            log(f"         trying filename override: '{clean}'")
        for pat in (f"*CREFC Financial File*{clean}*", f"*Financial*{clean}*"):
            hits = sorted(glob.glob(os.path.join(folder, pat)))
            log(f"         try: '{pat}' -> {[os.path.basename(h) for h in hits] if hits else 'no match'}")
            if hits:
                xlsx_hits = [h for h in hits if h.lower().endswith(".xlsx")]
                return (xlsx_hits or hits)[0]
    try:
        all_files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
        log(f"         folder contents ({len(all_files)} files): {all_files[:10]}")
    except Exception as e:
        log(f"         could not list folder: {e}")
    return ""


def xls_to_xlsx(src, dest, excel_app):
    """
    Convert an .xls file to .xlsx using a shared Excel COM instance.
    Preserves formatting by letting Excel handle the conversion.
    """
    ensure_dir(os.path.dirname(dest))
    if excel_app is None:
        raise RuntimeError("Excel COM instance is required to convert .xls to .xlsx")
    log(f"       Converting .xls -> .xlsx: {os.path.basename(src)} -> {os.path.basename(dest)}")
    try:
        wb = excel_app.Workbooks.Open(src)
        # 51 = xlOpenXMLWorkbook
        wb.SaveAs(dest, FileFormat=51)
        wb.Close(SaveChanges=False)
    except Exception as e:
        log_err(f"       Excel conversion failed for {src}: {e}")
        raise

# ── Generic .xls → .xlsx converter (no Excel COM required) ──────────────────

def convert_xls_to_xlsx(xls_path):
    """Convert .xls file to a temp .xlsx file. Returns path to the .xlsx copy."""
    import tempfile
    rb = xlrd.open_workbook(xls_path, formatting_info=False)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb_new = openpyxl.Workbook()
    for sheet_idx in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_idx)
        if sheet_idx == 0:
            ws_new = wb_new.active
            ws_new.title = rs.name
        else:
            ws_new = wb_new.create_sheet(rs.name)
        for row in range(rs.nrows):
            for col in range(rs.ncols):
                cell = rs.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    val = xlrd.xldate_as_datetime(cell.value, rb.datemode)
                ws_new.cell(row + 1, col + 1, val)
    wb_new.save(tmp.name)
    rb.release_resources()
    return tmp.name
# ── Periodic file ─────────────────────────────────────────────────────────────

def create_periodic(trans_id, det_date, irp_data, pirp_data, prev_folder, out_folder, excel_app=None):
    """
    Copy-forward previous month Periodic file.
    Updates 8 cells per loan row (matched by Loan ID in col C).
    All existing data rows are updated in-place — no rows added.
    """
    clean = clean_filename(trans_id)
    src   = find_file(prev_folder, "Periodic", trans_id)

    # TEST_MODE production fallback: test folder found but file missing
    if not src and cfg.TEST_MODE:
        prod_prev = prod_path(prev_folder)
        if prod_prev != prev_folder and os.path.isdir(prod_prev):
            src = find_file(prod_prev, "Periodic", trans_id)
            if src:
                log(f"     Periodic: Production fallback -> {os.path.basename(src)}")

    if not src:
        log(f"     Periodic: SKIP — no source file found in {prev_folder}")
        return "", "No previous Periodic file found", {}

    src_ext = os.path.splitext(src)[1].lower()
    dt      = parse_det_date(det_date)
    prefix  = dt.strftime("%Y.%m") if dt else det_date[:7]
    converted_tmp = None
    if src_ext == ".xls":
        converted_tmp = convert_xls_to_xlsx(src)
        src = converted_tmp
        log(f"     Converted .xls source to .xlsx for processing")
    dest    = os.path.join(out_folder, f"{prefix} CREFC Periodic File {clean}.xlsx")  # always .xlsx
    loan_map = get_loan_ids(irp_data, trans_id)
    prosp_map = build_prospectus_id_map(pirp_data, trans_id) if pirp_data else {}
    log(f"     Periodic: source={os.path.basename(src)}")
    log(f"     Periodic: dest={os.path.basename(dest)}")
    log(f"     Periodic: {len(loan_map)} IRP loan IDs to match")

    FIRST_DATA   = cfg.PERIODIC_FIRST_DATA - 1   # 0-based
    TRANS_COL    = cfg.PERIODIC_TRANS_COL  - 1   # 0-based
    LOAN_COL     = cfg.PERIODIC_LOAN_COL   - 1   # 0-based
    PROSP_COL_0  = 3   # col D (0-based) on .xls
    PROSP_COL_1  = 4   # col D (1-based) on .xlsx
    UPDATE_COLS = [c - 1 for c in cfg.PERIODIC_UPDATE_COLS]   # 0-based

    try:
        shutil.copy2(src, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        last_row = ws.max_row
        matched = unmatched = unmatched_logged = 0
        for row_1 in range(cfg.PERIODIC_FIRST_DATA, last_row + 1):
            raw_loan_id = ws.cell(row_1, cfg.PERIODIC_LOAN_COL).value
            if raw_loan_id is None or str(raw_loan_id).strip() == "":
                continue  # skip blank/separator rows
            irp_row = None
            variants = loan_id_variants(raw_loan_id)
            for key in variants:
                irp_row = loan_map.get(key)
                if irp_row:
                    break
            if irp_row:
                for col_1 in cfg.PERIODIC_UPDATE_COLS:
                    cell = ws.cell(row_1, col_1)
                    # Skip overwriting existing formulas to preserve template logic
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    cell.value = irp_data[irp_row][col_1]
                    # Col EE (135) = Reporting Period Begin Date: subtract 1 day.
                    # IRP reports day after det date (e.g. 12th); file needs det date (11th).
                    if col_1 == cfg.IRP_COL_RPT_BEGIN and cell.value is not None:
                        ee_before = cell.value
                        try:
                            v = cell.value
                            if isinstance(v, (datetime, date)):
                                cell.value = v - relativedelta(days=1)
                            else:
                                s = str(int(float(v)))[:8]
                                dt_ee = datetime.strptime(s, "%Y%m%d")
                                adjusted = dt_ee - relativedelta(days=1)
                                # Keep same type as original (int or float)
                                if isinstance(v, int):
                                    cell.value = int(adjusted.strftime("%Y%m%d"))
                                elif isinstance(v, float):
                                    cell.value = float(adjusted.strftime("%Y%m%d"))
                                else:
                                    cell.value = adjusted.strftime("%Y%m%d")
                        except Exception as e:
                            log(f"     Periodic EE date adjust failed row {row_1}: {e}", "WARN")
                    _highlight_cell(cell)

                # Prospectus Loan ID from PIRPXLLR (if available)
                prosp_loan_id = prosp_map.get(normalize_loan_id(raw_loan_id))
                if prosp_loan_id not in (None, ""):
                    try:
                        cell_prosp = ws.cell(row_1, PROSP_COL_1)
                        cell_prosp.value = prosp_loan_id
                        _highlight_cell(cell_prosp)
                    except Exception as e:
                        log(f"     Prospectus ID write failed row {row_1}: {e}", "WARN")

                # After IRP updates, ensure key Periodic formula columns
                # are present when cells are plain values:
                #
                # - G (col 7):  "Current Ending Scheduled Balance"  =F{row}-X{row}
                # - Y (col 25): "Total Scheduled P&I Due"          =W{row}+X{row}
                # - AJ (col 36): "Actual Balance"                  =F{row}-X{row}
                #
                # Only insert the formula when the cell does not already
                # contain a formula string.
                formula_cells = {
                    7:  f"=F{row_1}-X{row_1}",  # G
                    25: f"=W{row_1}+X{row_1}",  # Y
                    36: f"=F{row_1}-X{row_1}",  # AJ
                }
                for col_idx, formula in formula_cells.items():
                    cell = ws.cell(row_1, col_idx)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        _highlight_cell(cell)  # Still highlight — script owns this cell
                        continue
                    cell.value = formula
                    _highlight_cell(cell)

                matched += 1
            else:
                unmatched += 1
                if unmatched_logged < 5:
                    unmatched_logged += 1
                    log(
                        f"     Periodic (xlsx): unmatched row {row_1} "
                        f"raw='{raw_loan_id}' "
                        f"normalized='{normalize_loan_id(raw_loan_id)}' "
                        f"variants={variants}",
                        "WARN",
                    )
        wb.save(dest)

        metrics = {"matched": matched, "unmatched": unmatched, "total_rows": matched + unmatched}
        log(f"     Periodic: {matched} loans updated, {unmatched} unmatched (of {metrics['total_rows']} rows)")
        log(f"     Periodic: SAVED -> {dest}")
        for w in validate_output_file(dest, "Periodic"):
            log(f"     {w}", "WARN")
        return dest, "", metrics
    except Exception as e:
        return "", str(e), {}
    finally:
        if converted_tmp:
            try:
                os.unlink(converted_tmp)
            except Exception:
                pass

# ── Property file ─────────────────────────────────────────────────────────────

def create_property(trans_id, det_date, irp_data, prev_folder, out_folder, excel_app=None):
    """
    Copy-forward Property file.
    FIX GAP 1: Only updates the LAST BLOCK of rows (rows sharing same
    Distribution Date as the last row), not all rows in the file.
    Updates col E (Distribution Date) and col U (Allocated Balance).
    """
    clean = clean_filename(trans_id)
    src   = find_file(prev_folder, "Property", trans_id)

    if not src and cfg.TEST_MODE:
        prod_prev = prod_path(prev_folder)
        if prod_prev != prev_folder and os.path.isdir(prod_prev):
            src = find_file(prod_prev, "Property", trans_id)
            if src:
                log(f"     Property: Production fallback -> {os.path.basename(src)}")

    if not src:
        log(f"     Property: SKIP — no source file found in {prev_folder}")
        return "", "No previous Property file found", {}

    src_ext  = os.path.splitext(src)[1].lower()
    dt2      = parse_det_date(det_date)
    prefix2  = dt2.strftime("%Y.%m") if dt2 else det_date[:7]
    if src_ext == ".xls" and excel_app is not None:
        converted_src = os.path.join(out_folder, f"{prefix2} CREFC Property File {clean}_source.xlsx")
        xls_to_xlsx(src, converted_src, excel_app)
        src = converted_src
        src_ext = ".xlsx"
    dest     = os.path.join(out_folder, f"{prefix2} CREFC Property File {clean}.xlsx")
    loan_map = get_loan_ids(irp_data, trans_id)
    log(f"     Property: source={os.path.basename(src)}")
    log(f"     Property: dest={os.path.basename(dest)}")
    log(f"     Property: {len(loan_map)} IRP loan IDs to match")

    # 0-based constants for .xls path
    FIRST_DATA_0   = cfg.PROP_FIRST_DATA   - 1
    TRANS_COL_0    = cfg.PROP_TRANS_COL    - 1
    LOAN_COL_0     = cfg.PROP_LOAN_COL     - 1
    DIST_DATE_0    = cfg.PROP_DIST_DATE_COL - 1
    ALLOC_PCT_0    = cfg.PROP_ALLOC_PCT_COL - 1
    ALLOC_BAL_0    = cfg.PROP_ALLOC_BAL_COL - 1
    GANA_DIST_COL  = cfg.IRP_COL_DIST_DATE  # 1-based, used on irp_data
    GANA_END_COL   = cfg.IRP_COL_END_BAL

    try:
        if src_ext == ".xls":
            rb   = xlrd.open_workbook(src, formatting_info=True)
            wb_w = xl_copy(rb)
            ws_r = rb.sheet_by_index(0)
            ws_w = wb_w.get_sheet(0)

            # Find last data row
            last_row_0 = ws_r.nrows - 1
            while last_row_0 >= FIRST_DATA_0 and not xls_cell(ws_r, last_row_0, TRANS_COL_0):
                last_row_0 -= 1

            if last_row_0 < FIRST_DATA_0:
                log("     Property: No data rows in previous file - skipping")
                return "", "No data rows in previous Property file", {}

            log(f"     Property: Last data row: {last_row_0 + 1}")

            # FIX GAP 1: Detect block size by counting rows with same dist date
            last_dist = xls_cell(ws_r, last_row_0, DIST_DATE_0)
            block_size = 0
            for r in range(last_row_0, FIRST_DATA_0 - 1, -1):
                if str(xls_cell(ws_r, r, DIST_DATE_0) or "") == str(last_dist or ""):
                    block_size += 1
                else:
                    break
            if block_size == 0:
                block_size = last_row_0 - FIRST_DATA_0 + 1

            block_start_0 = last_row_0 - block_size + 1
            log(f"     Property: Block size={block_size}, updating rows {block_start_0+1}-{last_row_0+1}")

            matched = unmatched = unmatched_logged = 0
            alloc_written = 0
            seen_loan_ids = set()
            for row_0 in range(block_start_0, last_row_0 + 1):
                raw_loan_id = xls_cell(ws_r, row_0, LOAN_COL_0)
                if raw_loan_id is None or str(raw_loan_id).strip() == "":
                    continue  # skip blank/separator rows
                irp_row = None
                variants = list(loan_id_variants(raw_loan_id))
                for key in variants:
                    irp_row = loan_map.get(key)
                    if irp_row:
                        break
                if irp_row:
                    # E: Distribution Date — write on every row
                    ws_w.write(row_0, DIST_DATE_0, irp_data[irp_row][GANA_DIST_COL])
                    # U: Allocated Balance — only on first row per Loan ID
                    norm_lid = normalize_loan_id(raw_loan_id)
                    if norm_lid not in seen_loan_ids:
                        seen_loan_ids.add(norm_lid)
                        alloc_pct = 0.0
                        try:
                            alloc_pct = float(xls_cell(ws_r, row_0, ALLOC_PCT_0) or 0)
                        except Exception as e:
                            log(f"     Property (xls): alloc_pct parse failed row {row_0+1}: {e}", "WARN")
                        end_bal = 0.0
                        try:
                            end_bal = float(irp_data[irp_row][GANA_END_COL] or 0)
                        except Exception as e:
                            log(f"     Property (xls): end_bal parse failed row {row_0+1}: {e}", "WARN")
                        alloc_bal = end_bal * (alloc_pct / 100) if alloc_pct else end_bal
                        ws_w.write(row_0, ALLOC_BAL_0, alloc_bal)
                        alloc_written += 1
                    matched += 1
                else:
                    unmatched += 1
                    if unmatched_logged < 5:
                        unmatched_logged += 1
                        log(
                            f"     Property (xls): unmatched row {row_0+1} "
                            f"raw='{raw_loan_id}' "
                            f"normalized='{normalize_loan_id(raw_loan_id)}' "
                            f"variants={variants}",
                            "WARN",
                        )

            wb_w.save(dest)

        else:  # .xlsx
            shutil.copy2(src, dest)
            wb = openpyxl.load_workbook(dest)
            ws = wb.active

            # Find last data row
            last_row_1 = ws.max_row
            while last_row_1 >= cfg.PROP_FIRST_DATA and not ws.cell(last_row_1, cfg.PROP_TRANS_COL).value:
                last_row_1 -= 1

            if last_row_1 < cfg.PROP_FIRST_DATA:
                log("     Property: No data rows in previous file - skipping")
                wb.close()
                return "", "No data rows in previous Property file", {}

            # FIX GAP 1: Block detection
            last_dist = ws.cell(last_row_1, cfg.PROP_DIST_DATE_COL).value
            block_size = 0
            for r in range(last_row_1, cfg.PROP_FIRST_DATA - 1, -1):
                if str(ws.cell(r, cfg.PROP_DIST_DATE_COL).value or "") == str(last_dist or ""):
                    block_size += 1
                else:
                    break
            if block_size == 0:
                block_size = last_row_1 - cfg.PROP_FIRST_DATA + 1

            block_start_1 = last_row_1 - block_size + 1
            log(f"     Property: Block size={block_size}, updating rows {block_start_1}-{last_row_1}")

            matched = unmatched = unmatched_logged = 0
            alloc_written = 0
            seen_loan_ids = set()  # Track which loans already got col U written
            for row_1 in range(block_start_1, last_row_1 + 1):
                raw_loan_id = ws.cell(row_1, cfg.PROP_LOAN_COL).value
                if raw_loan_id is None or str(raw_loan_id).strip() == "":
                    continue  # skip blank/separator rows
                irp_row = None
                variants = list(loan_id_variants(raw_loan_id))
                for key in variants:
                    irp_row = loan_map.get(key)
                    if irp_row:
                        break
                if irp_row:
                    # Distribution Date — write on every row (same value, no harm)
                    cell_dist = ws.cell(row_1, cfg.PROP_DIST_DATE_COL)
                    cell_dist.value = irp_data[irp_row][GANA_DIST_COL]
                    _highlight_cell(cell_dist)

                    # Allocated Balance — only on FIRST row per Loan ID
                    norm_lid = normalize_loan_id(raw_loan_id)
                    if norm_lid not in seen_loan_ids:
                        seen_loan_ids.add(norm_lid)
                        alloc_pct = 0.0
                        try:
                            alloc_pct = float(ws.cell(row_1, cfg.PROP_ALLOC_PCT_COL).value or 0)
                        except Exception as e:
                            log(f"     Property (xlsx): alloc_pct parse failed row {row_1}: {e}", "WARN")
                        end_bal = 0.0
                        try:
                            end_bal = float(irp_data[irp_row][GANA_END_COL] or 0)
                        except Exception as e:
                            log(f"     Property (xlsx): end_bal parse failed row {row_1}: {e}", "WARN")
                        cell_bal = ws.cell(row_1, cfg.PROP_ALLOC_BAL_COL)
                        cell_bal.value = end_bal * (alloc_pct / 100) if alloc_pct else end_bal
                        _highlight_cell(cell_bal)
                        alloc_written += 1
                    matched += 1
                else:
                    unmatched += 1
                    if unmatched_logged < 5:
                        unmatched_logged += 1
                        log(
                            f"     Property (xlsx): unmatched row {row_1} "
                            f"raw='{raw_loan_id}' "
                            f"normalized='{normalize_loan_id(raw_loan_id)}' "
                            f"variants={variants}",
                            "WARN",
                        )
            wb.save(dest)

        metrics = {"matched": matched, "unmatched": unmatched, "total_rows": matched + unmatched}
        log(f"     Property: {matched} rows updated, {unmatched} unmatched (of {metrics['total_rows']} rows)")
        log(f"     Property: SAVED -> {dest}")
        for w in validate_output_file(dest, "Property"):
            log(f"     {w}", "WARN")
        return dest, "", metrics
    except Exception as e:
        return "", str(e), {}

# ── Supplemental file ─────────────────────────────────────────────────────────

def _supp_a1_update_xls(ws_w, ws_r, new_date_str, tab_name):
    """
    FIX GAP 2: Replace just the date in A1, not overwrite the whole cell.
    Matches VBA UpdateSupplementalA1Date exactly.
    """
    a1_val = str(xls_cell(ws_r, 0, 0) or "")
    if not a1_val:
        log(f"       {tab_name}: A1 is empty - skipped")
        return
    updated = replace_date_in_text(a1_val, new_date_str)
    if updated != a1_val:
        ws_w.write(0, 0, updated)
        log(f"       {tab_name}: A1 date updated to {new_date_str}")
    else:
        log(f"       {tab_name}: No date pattern found in A1 - skipped")

def _supp_a1_update_xlsx(ws, new_date_str, tab_name):
    """FIX GAP 2 (.xlsx path): same replace_date_in_text logic."""
    a1_val = str(ws["A1"].value or "")
    if not a1_val:
        log(f"       {tab_name}: A1 is empty - skipped")
        return
    updated = replace_date_in_text(a1_val, new_date_str)
    if updated != a1_val:
        ws["A1"] = updated
        _highlight_cell(ws["A1"])
        log(f"       {tab_name}: A1 date updated to {new_date_str}")
    else:
        log(f"       {tab_name}: No date pattern found in A1 - skipped")

def _parse_formula_cell_ref(formula_str):
    """
    Parse a formula that references another cell, e.g. ='Comp Finan Status'!B14 or =Sheet!A1.
    Returns (sheet_name, cell_ref) or None if not a simple sheet!cell reference.
    """
    if not formula_str or not isinstance(formula_str, str):
        return None
    s = formula_str.strip()
    if not s.startswith("="):
        return None
    s = s[1:].strip()
    # Quoted sheet: 'Sheet Name'!A1 or '[1]Sheet Name'!A1 (external workbook ref)
    m = re.match(r"^'([^']+)'\s*!\s*(.+)$", s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # Unquoted sheet: SheetName!A1
    m = re.match(r"^([^!]+)\s*!\s*(.+)$", s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None

def _normalize_formula_sheet_name(sheet_name):
    """Strip external workbook prefix like [1] or [2] from parsed sheet name."""
    if not sheet_name or not isinstance(sheet_name, str):
        return (sheet_name or "").strip()
    return re.sub(r"^\[\d+\]\s*", "", sheet_name.strip()).strip()

def _fuzzy_sheet_match(formula_sheet_name, actual_sheet_name):
    """
    True if each word in formula_sheet_name is a prefix of the corresponding word in actual_sheet_name.
    E.g. 'Comp Finan Status' matches 'Comparative Financial Status' (Comp→Comparative, Finan→Financial, Status→Status).
    """
    if not formula_sheet_name or not actual_sheet_name:
        return False
    f_words = formula_sheet_name.split()
    a_words = actual_sheet_name.split()
    if len(f_words) != len(a_words):
        return False
    return all(a_words[i].lower().startswith(f_words[i].lower()) for i in range(len(f_words)))

def _find_sheet_name(sheet_names, formula_sheet_name):
    """
    Resolve formula sheet name (e.g. 'Comp Finan Status') to an actual tab name in the workbook.
    Tries exact match first, then fuzzy word-prefix match (abbrev vs long-form, e.g. Comparative Financial Status).
    Returns the actual sheet name or None.
    """
    normalized = _normalize_formula_sheet_name(formula_sheet_name)
    if not normalized:
        return None
    if normalized in sheet_names:
        return normalized
    for name in sheet_names:
        if _fuzzy_sheet_match(normalized, name):
            return name
    return None

def _a1_to_xlrd_rc(cell_ref):
    """Convert A1-style cell ref to xlrd (row_0, col_0). E.g. 'B14' or '$B$14' -> (13, 1)."""
    clean = (cell_ref or "").strip().replace("$", "")
    m = re.match(r"^([A-Za-z]+)(\d+)$", clean)
    if not m:
        return None
    col_letters, row_str = m.group(1), m.group(2)
    try:
        # openpyxl column_index_from_string is 1-based (A=1, B=2)
        col_1 = openpyxl.utils.column_index_from_string(col_letters.upper())
        col_0 = col_1 - 1
        row_0 = int(row_str) - 1
        return row_0, col_0
    except Exception:
        return None

def _resolve_loan_id_from_formula_xlsx(wb, value):
    """If value is a formula like ='Sheet'!A1, return that cell's value; else return value."""
    parsed = _parse_formula_cell_ref(value)
    if not parsed:
        return value
    formula_sheet_name, cell_ref = parsed
    try:
        sheet_name = _find_sheet_name(wb.sheetnames, formula_sheet_name)
        if not sheet_name:
            return value
        ws_ref = wb[sheet_name]
        # openpyxl expects A1-style without $ for indexing
        a1 = (cell_ref or "").replace("$", "")
        return ws_ref[a1].value
    except Exception:
        return value

def _resolve_loan_id_from_formula_xls(book_r, value):
    """If value is a formula like ='Sheet'!A1, return that cell's value from xlrd book; else return value."""
    parsed = _parse_formula_cell_ref(value)
    if not parsed:
        return value
    formula_sheet_name, cell_ref = parsed
    try:
        sheet_names = [book_r.sheet_by_index(i).name for i in range(book_r.nsheets)]
        sheet_name = _find_sheet_name(sheet_names, formula_sheet_name)
        if not sheet_name:
            return value
        sheet = book_r.sheet_by_name(sheet_name)
        rc = _a1_to_xlrd_rc(cell_ref)
        if rc is None:
            return value
        row_0, col_0 = rc
        return sheet.cell_value(row_0, col_0)
    except Exception:
        return value

def _find_total_row_xls(ws_r, first_data_row_0):
    """Scan col A for a cell containing 'TOTAL' (case-insensitive)."""
    for r in range(first_data_row_0, ws_r.nrows):
        v = str(xls_cell(ws_r, r, 0) or "").upper()
        if "TOTAL" in v:
            return r
    return -1

def _find_total_row_xlsx(ws, first_data_row_1):
    """Scan col A for 'TOTAL' text."""
    for r in range(first_data_row_1, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or "").upper()
        if "TOTAL" in v:
            return r
    return -1

def _block_size_xls(ws_r, last_data_row_0, first_data_row_0, key_col_0):
    """Count consecutive non-empty rows from last_data_row backwards."""
    block = 0
    for r in range(last_data_row_0, first_data_row_0 - 1, -1):
        if str(xls_cell(ws_r, r, key_col_0) or "").strip():
            block += 1
        else:
            break
    return block if block > 0 else 1

def _block_size_xlsx(ws, last_data_row_1, first_data_row_1, key_col_1):
    block = 0
    for r in range(last_data_row_1, first_data_row_1 - 1, -1):
        if str(ws.cell(r, key_col_1).value or "").strip():
            block += 1
        else:
            break
    return block if block > 0 else 1

def _process_total_loan_xls(ws_r, ws_w, loan_map, irp_data, trans_id, book_r):
    """
    Total Loan tab.
    Key col = C (index 2, 0-based). First data row = 11 (index 10, 0-based).
    If key cell contains a formula (e.g. ='Comp Finan Status'!B14), it is resolved from book_r.
    Columns updated per reference file DBJPM 2017-C6:
      L (0-based 11) = Current Ending Scheduled Balance  (IRP col 7)  ← v2.2 addition
      M (0-based 12) = Total Scheduled P&I Due           (IRP col 25)
      O (0-based 14) = Paid Through Date                 (IRP col 8, YYYYMMDD float)
    """
    FIRST_DATA_0  = 10   # row 11 in 1-based
    KEY_COL_0     = 2    # col C  (0-based)
    END_BAL_COL_0 = 11   # col L  (0-based 11 = 1-based 12)  ← v2.2
    PI_COL_0      = 12   # col M  (0-based 12 = 1-based 13)
    PTD_COL_0     = 14   # col O  (0-based 14 = 1-based 15)

    total_row_0 = _find_total_row_xls(ws_r, FIRST_DATA_0)
    if total_row_0 < 0:
        log("       Total Loan: TOTAL row not found - skipping")
        return

    # Last data row = row before TOTAL with non-empty key
    last_data_0 = total_row_0 - 1
    while last_data_0 >= FIRST_DATA_0 and not str(xls_cell(ws_r, last_data_0, KEY_COL_0) or "").strip():
        last_data_0 -= 1
    if last_data_0 < FIRST_DATA_0:
        log("       Total Loan: No data rows found")
        return

    block = _block_size_xls(ws_r, last_data_0, FIRST_DATA_0, KEY_COL_0)
    block_start = last_data_0 - block + 1
    log(f"       Total Loan: Updating rows {block_start+1}-{last_data_0+1} ({block} rows)")

    matched = unmatched = unmatched_logged = 0
    for row_0 in range(block_start, last_data_0 + 1):
        raw_loan_id = xls_cell(ws_r, row_0, KEY_COL_0)
        raw_loan_id = _resolve_loan_id_from_formula_xls(book_r, raw_loan_id)
        irp_row = None
        variants = list(loan_id_variants(raw_loan_id))
        for key in variants:
            irp_row = loan_map.get(key)
            if irp_row:
                break
        if irp_row:
            ws_w.write(row_0, END_BAL_COL_0, irp_data[irp_row][cfg.IRP_COL_END_BAL])   # v2.2
            ws_w.write(row_0, PI_COL_0,      irp_data[irp_row][cfg.SUPP_IRP_PI_COL])
            ws_w.write(row_0, PTD_COL_0,     irp_date_to_yyyymmdd(irp_data[irp_row][cfg.IRP_COL_PAID_THRU]))
            matched += 1
        else:
            unmatched += 1
            if unmatched_logged < 5:
                unmatched_logged += 1
                log(
                    f"       Total Loan (xls): unmatched row {row_0+1} "
                    f"raw='{raw_loan_id}' "
                    f"normalized='{normalize_loan_id(raw_loan_id)}' "
                    f"variants={variants}",
                    "WARN",
                )
    log(f"       Total Loan: {matched} of {block} rows updated (L=EndBal, M=P&I, O=PTD; {unmatched} unmatched)")

    # Ensure Total row formulas on row 16 (E/J/K/L/M) are present.
    # TOTAL row is where col A contains 'TOTAL'.
    total_row_0 = _find_total_row_xls(ws_r, FIRST_DATA_0)
    if total_row_0 >= 0:
        excel_total_row = total_row_0 + 1
        first_data_row_1 = FIRST_DATA_0 + 1  # 1-based row of first data row (usually 11)
        # Static ranges per requirements: E11:E15, J11:J15, K11:K15, L11:L15, M11:M15
        try:
            # Preserve styles for total row cells when possible.
            def _style(col_idx):
                try:
                    return ws_w.cell(total_row_0, col_idx).xf
                except Exception:
                    return None

            style_E = _style(4)
            style_J = _style(9)
            style_K = _style(10)
            style_L = _style(11)
            style_M = _style(12)

            def _write(col_idx, formula, style):
                if style is not None:
                    ws_w.write(total_row_0, col_idx, xlwt.Formula(formula), style)
                else:
                    ws_w.write(total_row_0, col_idx, xlwt.Formula(formula))

            _write(4,  f"SUM(E{first_data_row_1}:E{excel_total_row-1})", style_E)  # E
            _write(9,  f"SUM(J{first_data_row_1}:J{excel_total_row-1})", style_J)  # J
            _write(10, f"SUM(K{first_data_row_1}:K{excel_total_row-1})", style_K)  # K
            _write(11, f"SUM(L{first_data_row_1}:L{excel_total_row-1})", style_L)  # L
            _write(12, f"SUM(M{first_data_row_1}:M{excel_total_row-1})", style_M)  # M
        except Exception as e:
            log(f"     Total Loan (xls): TOTAL row formula write failed: {e}", "WARN")

def _process_total_loan_xlsx(ws, loan_map, irp_data):
    """
    Total Loan tab (.xlsx path).
    If key col C contains a formula (e.g. ='Comp Finan Status'!B14), it is resolved from the workbook.
    L (col 12) = Current Ending Scheduled Balance  (IRP col 7)  ← v2.2
    M (col 13) = Total Scheduled P&I Due           (IRP col 25)
    O (col 15) = Paid Through Date                 (IRP col 8, YYYYMMDD float)
    """
    FIRST_DATA_1  = 11
    KEY_COL_1     = 3    # col C (1-based)
    END_BAL_COL_1 = 12   # col L  ← v2.2
    PI_COL_1      = 13   # col M
    PTD_COL_1     = 15   # col O

    total_row_1 = _find_total_row_xlsx(ws, FIRST_DATA_1)
    if total_row_1 < 0:
        log("       Total Loan: TOTAL row not found - skipping")
        return

    last_data_1 = total_row_1 - 1
    while last_data_1 >= FIRST_DATA_1 and not str(ws.cell(last_data_1, KEY_COL_1).value or "").strip():
        last_data_1 -= 1
    if last_data_1 < FIRST_DATA_1:
        log("       Total Loan: No data rows found")
        return

    block = _block_size_xlsx(ws, last_data_1, FIRST_DATA_1, KEY_COL_1)
    block_start = last_data_1 - block + 1
    wb = ws.parent
    matched = unmatched = unmatched_logged = 0
    for row_1 in range(block_start, last_data_1 + 1):
        raw_loan_id = ws.cell(row_1, KEY_COL_1).value
        raw_loan_id = _resolve_loan_id_from_formula_xlsx(wb, raw_loan_id)
        irp_row = None
        variants = list(loan_id_variants(raw_loan_id))
        for key in variants:
            irp_row = loan_map.get(key)
            if irp_row:
                break
        if irp_row:
            cell_L = ws.cell(row_1, END_BAL_COL_1)
            cell_L.value = irp_data[irp_row][cfg.IRP_COL_END_BAL]   # v2.2
            _highlight_cell(cell_L)
            cell_M = ws.cell(row_1, PI_COL_1)
            cell_M.value = irp_data[irp_row][cfg.SUPP_IRP_PI_COL]
            _highlight_cell(cell_M)
            cell_O = ws.cell(row_1, PTD_COL_1)
            cell_O.value = irp_date_to_yyyymmdd(irp_data[irp_row][cfg.IRP_COL_PAID_THRU])
            _highlight_cell(cell_O)
            matched += 1
        else:
            unmatched += 1
            if unmatched_logged < 5:
                unmatched_logged += 1
                log(
                    f"       Total Loan (xlsx): unmatched row {row_1} "
                    f"raw='{raw_loan_id}' "
                    f"normalized='{normalize_loan_id(raw_loan_id)}' "
                    f"variants={variants}",
                    "WARN",
                )
    log(f"       Total Loan: {matched} of {block} rows updated (L=EndBal, M=P&I, O=PTD; {unmatched} unmatched)")

    # Ensure Total row formulas on row 16 (E/J/K/L/M) are present.
    total_row_1 = _find_total_row_xlsx(ws, FIRST_DATA_1)
    if total_row_1 >= 0:
        first_data_row_1 = FIRST_DATA_1  # usually 11
        try:
            for _tc, _tf in [
                (5,  f"=SUM(E{first_data_row_1}:E{total_row_1-1})"),
                (10, f"=SUM(J{first_data_row_1}:J{total_row_1-1})"),
                (11, f"=SUM(K{first_data_row_1}:K{total_row_1-1})"),
                (12, f"=SUM(L{first_data_row_1}:L{total_row_1-1})"),
                (13, f"=SUM(M{first_data_row_1}:M{total_row_1-1})"),
            ]:
                _cell = ws.cell(total_row_1, _tc)
                _cell.value = _tf
                _highlight_cell(_cell)
        except Exception as e:
            log(f"     Total Loan (xlsx): TOTAL row formula write failed: {e}", "WARN")

def _process_comp_finan_xls(ws_r, ws_w, loan_map, irp_data):
    """
    FIX GAP 3: Key col = B (index 1, not 2). Block-only update.
    Updates col I (Ending Balance) and col J (Paid Through Date).
    First data row = 14 (index 13).
    """
    FIRST_DATA_0 = 13   # row 14
    KEY_COL_0    = 1    # col B  ← FIX: was 2 (col C), correct is 1 (col B)
    UPDATE_I_0   = 8    # col I
    UPDATE_J_0   = 9    # col J

    total_row_0 = _find_total_row_xls(ws_r, FIRST_DATA_0)
    if total_row_0 < 0:
        log("       Comp Finan: TOTAL row not found - skipping")
        return

    last_data_0 = total_row_0 - 1
    while last_data_0 >= FIRST_DATA_0 and not str(xls_cell(ws_r, last_data_0, KEY_COL_0) or "").strip():
        last_data_0 -= 1
    if last_data_0 < FIRST_DATA_0:
        log("       Comp Finan: No data rows found")
        return

    block = _block_size_xls(ws_r, last_data_0, FIRST_DATA_0, KEY_COL_0)
    block_start = last_data_0 - block + 1
    log(f"       Comp Finan: Updating rows {block_start+1}-{last_data_0+1} ({block} rows)")

    matched = unmatched = unmatched_logged = 0
    ij_written = 0
    seen_loan_ids = set()
    for row_0 in range(block_start, last_data_0 + 1):
        raw_loan_id = xls_cell(ws_r, row_0, KEY_COL_0)
        irp_row = None
        variants = list(loan_id_variants(raw_loan_id))
        for key in variants:
            irp_row = loan_map.get(key)
            if irp_row:
                break
        if irp_row:
            norm_lid = normalize_loan_id(raw_loan_id)
            if norm_lid not in seen_loan_ids:
                seen_loan_ids.add(norm_lid)
                ws_w.write(row_0, UPDATE_I_0, irp_data[irp_row][cfg.IRP_COL_END_BAL])
                ws_w.write(row_0, UPDATE_J_0, irp_data[irp_row][cfg.IRP_COL_PAID_THRU])
                ij_written += 1
            matched += 1
        else:
            unmatched += 1
            if unmatched_logged < 5:
                unmatched_logged += 1
                log(
                    f"       Comp Finan (xls): unmatched row {row_0+1} "
                    f"raw='{raw_loan_id}' "
                    f"normalized='{normalize_loan_id(raw_loan_id)}' "
                    f"variants={variants}",
                    "WARN",
                )
    log(f"       Comp Finan: {matched} of {block} rows updated ({unmatched} unmatched)")

    # Insert Comp Finan Status total formulas — dynamic range
    total_row_0 = _find_total_row_xls(ws_r, FIRST_DATA_0)
    if total_row_0 >= 0:
        last_data_for_total = total_row_0  # 0-based, so this is the row BEFORE total in 1-based
        first_data_1based = FIRST_DATA_0 + 1  # convert 0-based first data to 1-based
        try:
            def _style(col_idx):
                try:
                    return ws_w.cell(total_row_0, col_idx).xf
                except Exception:
                    return None

            formula_cols = [
                (8,  "I"),
                (12, "M"),
                (13, "N"),
                (17, "R"),
                (18, "S"),
                (22, "W"),
                (23, "X"),
                (29, "AD"),
                (30, "AE"),
                (33, "AH"),
            ]

            for col_idx, col_letter in formula_cols:
                style = _style(col_idx)
                formula = f"SUM({col_letter}{first_data_1based}:{col_letter}{last_data_for_total})"
                if style is not None:
                    ws_w.write(total_row_0, col_idx, xlwt.Formula(formula), style)
                else:
                    ws_w.write(total_row_0, col_idx, xlwt.Formula(formula))
        except Exception:
            pass

def _process_comp_finan_xlsx(ws, loan_map, irp_data):
    """FIX GAP 3 (.xlsx path): Key col = B (col 2 in 1-based openpyxl)."""
    FIRST_DATA_1 = 14
    KEY_COL_1    = 2    # col B  ← FIX
    UPDATE_I_1   = 9    # col I
    UPDATE_J_1   = 10   # col J

    total_row_1 = _find_total_row_xlsx(ws, FIRST_DATA_1)
    if total_row_1 < 0:
        log("       Comp Finan: TOTAL row not found - skipping")
        return

    last_data_1 = total_row_1 - 1
    while last_data_1 >= FIRST_DATA_1 and not str(ws.cell(last_data_1, KEY_COL_1).value or "").strip():
        last_data_1 -= 1
    if last_data_1 < FIRST_DATA_1:
        log("       Comp Finan: No data rows found")
        return

    block = _block_size_xlsx(ws, last_data_1, FIRST_DATA_1, KEY_COL_1)
    block_start = last_data_1 - block + 1
    matched = unmatched = unmatched_logged = 0
    ij_written = 0
    seen_loan_ids = set()  # Only update I & J on first row per loan
    for row_1 in range(block_start, last_data_1 + 1):
        raw_loan_id = ws.cell(row_1, KEY_COL_1).value
        irp_row = None
        variants = list(loan_id_variants(raw_loan_id))
        for key in variants:
            irp_row = loan_map.get(key)
            if irp_row:
                break
        if irp_row:
            norm_lid = normalize_loan_id(raw_loan_id)
            if norm_lid not in seen_loan_ids:
                seen_loan_ids.add(norm_lid)
                cell_I = ws.cell(row_1, UPDATE_I_1)
                cell_I.value = irp_data[irp_row][cfg.IRP_COL_END_BAL]
                _highlight_cell(cell_I)
                cell_J = ws.cell(row_1, UPDATE_J_1)
                cell_J.value = irp_data[irp_row][cfg.IRP_COL_PAID_THRU]
                _highlight_cell(cell_J)
                ij_written += 1
            matched += 1
        else:
            unmatched += 1
            if unmatched_logged < 5:
                unmatched_logged += 1
                log(
                    f"       Comp Finan (xlsx): unmatched row {row_1} "
                    f"raw='{raw_loan_id}' "
                    f"normalized='{normalize_loan_id(raw_loan_id)}' "
                    f"variants={variants}",
                    "WARN",
                )
    log(f"       Comp Finan: {matched} of {block} rows updated ({unmatched} unmatched)")

    # Insert Comp Finan Status total formulas — dynamic range
    total_row_1 = _find_total_row_xlsx(ws, FIRST_DATA_1)
    if total_row_1 >= 0:
        last_data_for_total = total_row_1 - 1
        try:
            for _tc, _col_letter in [
                (9,  "I"),
                (13, "M"),
                (14, "N"),
                (18, "R"),
                (19, "S"),
                (23, "W"),
                (24, "X"),
                (30, "AD"),
                (31, "AE"),
                (34, "AH"),
            ]:
                _cell = ws.cell(total_row_1, _tc)
                _cell.value = f"=SUM({_col_letter}{FIRST_DATA_1}:{_col_letter}{last_data_for_total})"
                _highlight_cell(_cell)
        except Exception:
            pass

def _collect_res_loc_rows(pirp_data, trans_id):
    """
    FIX GAP 4: Track currentTransID across rows because PIRPXLLR col A
    only has the Transaction ID on the FIRST row of each loan group.
    Continuation rows (same loan, different reserve type) have blank col A.
    Copies 15 columns (A:O), matching the VBA RL_LAST_COL = 15.
    """
    # Find where data starts: scan for header row where col A AND col B = "ID"
    data_start = 7   # default (row 8 in 1-based = index 7 in 0-based, but pirp_data is 1-padded)
    for i, row in enumerate(pirp_data[1:], start=1):
        if (str(row[1] or "").upper().strip() == "ID" and
                str(row[2] or "").upper().strip() == "ID"):
            data_start = i + 1
            break

    # Find last data row (col C = Loan ID most consistently filled)
    last_data = len(pirp_data) - 1
    while last_data >= data_start and not str(pirp_data[last_data][3] or "").strip():
        last_data -= 1

    # Stop before any TOTALS row
    for i in range(data_start, last_data + 1):
        if str(pirp_data[i][1] or "").upper().strip() == "TOTALS":
            last_data = i - 1
            break

    log(f"       Res LOC: PIRPXLLR data rows {data_start} to {last_data}")

    # Collect matching rows, tracking currentTransID across blank-col-A rows
    matching = []
    current_tid = ""
    for i in range(data_start, last_data + 1):
        col_a = str(pirp_data[i][1] or "").strip()
        if col_a and col_a != "0":
            current_tid = col_a
        if current_tid.upper() == trans_id.upper():
            matching.append(i)

    log(f"       Res LOC: {len(matching)} matching rows from PIRPXLLR")
    return matching

def _process_res_loc_xls(ws_r, ws_w, pirp_data, trans_id):
    """
    FIX GAP 4: Use continuation-row-aware collection, 15 cols (A:O).
    Overwrites from row 8 (index 7) downward. Fills Trans ID on every row.
    """
    WRITE_START_0 = 7    # row 8 in 1-based = index 7
    LAST_COL      = 15   # A:O

    matching = _collect_res_loc_rows(pirp_data, trans_id)
    if not matching:
        log(f"       Res LOC: No matching data in PIRPXLLR for {trans_id}")
        return

    for i, pirp_row_idx in enumerate(matching):
        write_0 = WRITE_START_0 + i
        for c in range(1, LAST_COL + 1):
            val = pirp_data[pirp_row_idx][c] if c < len(pirp_data[pirp_row_idx]) else None
            if val is not None:
                ws_w.write(write_0, c - 1, val)
        # Ensure Transaction ID on every row
        if not str(xls_cell(ws_r, write_0, 0) or "").strip():
            ws_w.write(write_0, 0, trans_id)

    log(f"       Res LOC: {len(matching)} rows written from PIRPXLLR")

    last_written_0 = WRITE_START_0 + len(matching) - 1  # last row with data (0-based)

    cleared = 0
    # Find the actual TOTALS row by scanning col A
    total_row_0 = _find_total_row_xls(ws_r, WRITE_START_0)
    if total_row_0 < 0:
        log("       Res LOC: TOTALS row not found - skipping formulas")
    else:
        # Clear stale rows between last written data and TOTALS
        for stale_r in range(last_written_0 + 1, total_row_0):
            row_has_data = False
            for c in range(0, LAST_COL):
                if xls_cell(ws_r, stale_r, c) is not None and str(xls_cell(ws_r, stale_r, c)).strip():
                    row_has_data = True
                    ws_w.write(stale_r, c, "")
            if row_has_data:
                cleared += 1
        if cleared:
            log(f"       Res LOC: Cleared {cleared} stale rows ({last_written_0 + 2} to {total_row_0})")

        # Write TOTALS formulas with dynamic range
        first_1 = WRITE_START_0 + 1  # 1-based row of first data row
        last_1 = last_written_0 + 1  # 1-based row of last data row
        try:
            def _style(col_idx):
                try:
                    return ws_w.cell(total_row_0, col_idx).xf
                except Exception:
                    return None

            for col_idx, col_letter in [
                (8,  "I"),
                (9,  "J"),
                (10, "K"),
                (11, "L"),
                (12, "M"),
            ]:
                style = _style(col_idx)
                formula = f"SUM({col_letter}{first_1}:{col_letter}{last_1})"
                if style is not None:
                    ws_w.write(total_row_0, col_idx, xlwt.Formula(formula), style)
                else:
                    ws_w.write(total_row_0, col_idx, xlwt.Formula(formula))

            # M column formula on data rows only
            for r_0 in range(WRITE_START_0, last_written_0 + 1):
                r_1 = r_0 + 1  # 1-based for Excel formula
                try:
                    style_m = ws_w.cell(r_0, 12).xf
                except Exception:
                    style_m = None
                if style_m is not None:
                    ws_w.write(r_0, 12, xlwt.Formula(f"J{r_1}+K{r_1}-L{r_1}"), style_m)
                else:
                    ws_w.write(r_0, 12, xlwt.Formula(f"J{r_1}+K{r_1}-L{r_1}"))
        except Exception:
            pass

def _process_res_loc_xlsx(ws, pirp_data, trans_id):
    """FIX GAP 4 (.xlsx path)."""
    WRITE_START_1 = 8
    LAST_COL      = 15

    matching = _collect_res_loc_rows(pirp_data, trans_id)
    if not matching:
        log(f"       Res LOC: No matching data in PIRPXLLR for {trans_id}")
        return

    for i, pirp_row_idx in enumerate(matching):
        write_1 = WRITE_START_1 + i
        for c in range(1, LAST_COL + 1):
            val = pirp_data[pirp_row_idx][c] if c < len(pirp_data[pirp_row_idx]) else None
            if val is not None:
                cell_rl = ws.cell(write_1, c)
                cell_rl.value = val
                _highlight_cell(cell_rl)
        if not str(ws.cell(write_1, 1).value or "").strip():
            cell_tid = ws.cell(write_1, 1)
            cell_tid.value = trans_id
            _highlight_cell(cell_tid)

    log(f"       Res LOC: {len(matching)} rows written from PIRPXLLR")

    last_written_row = WRITE_START_1 + len(matching) - 1  # last row with actual data

    cleared = 0
    # Find the actual TOTALS row by scanning col A
    total_row_1 = _find_total_row_xlsx(ws, WRITE_START_1)
    if total_row_1 < 0:
        # No TOTALS row found — skip formula insertion
        log("       Res LOC: TOTALS row not found - skipping formulas")
    else:
        # Clear stale rows between last written data and TOTALS
        for stale_r in range(last_written_row + 1, total_row_1):
            row_has_data = False
            for c in range(1, LAST_COL + 1):
                if ws.cell(stale_r, c).value is not None:
                    row_has_data = True
                    ws.cell(stale_r, c).value = None
            if row_has_data:
                cleared += 1
        if cleared:
            log(f"       Res LOC: Cleared {cleared} stale rows ({last_written_row + 1} to {total_row_1 - 1})")

        # Write TOTALS formulas with dynamic range
        try:
            for _tc, _col_letter in [
                (9,  "I"),
                (10, "J"),
                (11, "K"),
                (12, "L"),
                (13, "M"),
            ]:
                _cell = ws.cell(total_row_1, _tc)
                _cell.value = f"=SUM({_col_letter}{WRITE_START_1}:{_col_letter}{last_written_row})"
                _highlight_cell(_cell)

            # M column formula (Ending Reserve = Begin + Deposits - Disbursements)
            for r in range(WRITE_START_1, last_written_row + 1):
                _cell_m = ws.cell(r, 13)
                _cell_m.value = f"=J{r}+K{r}-L{r}"
                _highlight_cell(_cell_m)
        except Exception:
            pass

def create_supplemental(trans_id, det_date, irp_data, pirp_data, prev_folder, out_folder, excel_app=None):
    """
    Copy-forward Supplemental file (7 tabs).
    FIX GAP 2: A1 date replacement instead of overwrite.
    FIX GAP 3: Comp Finan uses col B key + block detection.
    FIX GAP 4: Res LOC uses continuation-row logic + 15 cols.
    """
    clean = clean_filename(trans_id)
    src   = find_file(prev_folder, "Supplemental", trans_id)

    if not src and cfg.TEST_MODE:
        prod_prev = prod_path(prev_folder)
        if prod_prev != prev_folder and os.path.isdir(prod_prev):
            src = find_file(prod_prev, "Supplemental", trans_id)
            if src:
                log(f"     Supplemental: Production fallback -> {os.path.basename(src)}")

    if not src:
        log(f"     Supplemental: SKIP — no source file found in {prev_folder}")
        return "", "No previous Supplemental file found", {}

    src_ext  = os.path.splitext(src)[1].lower()
    dt3      = parse_det_date(det_date)
    prefix3  = dt3.strftime("%Y.%m") if dt3 else det_date[:7]
    if src_ext == ".xls" and excel_app is not None:
        converted_src = os.path.join(out_folder, f"{prefix3} CREFC Supplemental {clean}_source.xlsx")
        xls_to_xlsx(src, converted_src, excel_app)
        src = converted_src
        src_ext = ".xlsx"
    dest     = os.path.join(out_folder, f"{prefix3} CREFC Supplemental {clean}.xlsx")
    loan_map = get_loan_ids(irp_data, trans_id)
    log(f"     Supplemental: source={os.path.basename(src)}")
    log(f"     Supplemental: dest={os.path.basename(dest)}")
    log(f"     Supplemental: {len(loan_map)} IRP loan IDs to match")

    # Date string: mm/dd/yy (2-digit year) matching VBA Format(dt, "mm/dd/yy")
    dt = parse_det_date(det_date)
    new_date_str = format_date_short(dt) if dt else det_date

    log(f"     Supplemental: Source file: {os.path.basename(src)}")

    try:
        if src_ext == ".xls":
            rb   = xlrd.open_workbook(src, formatting_info=True)
            wb_w = xl_copy(rb)

            tab_names = [rb.sheet_by_index(i).name for i in range(rb.nsheets)]
            log(f"     Supplemental: {rb.nsheets} tabs found: {tab_names}")
            for sheet_idx in range(rb.nsheets):
                ws_r   = rb.sheet_by_index(sheet_idx)
                ws_w   = wb_w.get_sheet(sheet_idx)
                sname  = ws_r.name

                # FIX GAP 2: All tabs get date-in-text replacement
                _supp_a1_update_xls(ws_w, ws_r, new_date_str, sname)

                if sname == "Total Loan":
                    log(f"       Processing Total Loan tab...")
                    _process_total_loan_xls(ws_r, ws_w, loan_map, irp_data, trans_id, rb)
                elif sname == "Comp Finan Status":
                    log(f"       Processing Comp Finan Status tab...")
                    _process_comp_finan_xls(ws_r, ws_w, loan_map, irp_data)
                elif sname == "Res LOC Report":
                    log(f"       Processing Res LOC Report tab...")
                    if pirp_data:
                        _process_res_loc_xls(ws_r, ws_w, pirp_data, trans_id)
                    else:
                        log("       Res LOC: No PIRPXLLR data - carried forward as-is")
                elif sname == "Watchlist":
                    # Watchlist: A1 date update only.
                    # Watchlist data is managed by the servicer — do not inject formulas.
                    log(f"       Tab 'Watchlist': A1 date update only (no formula injection)")
                else:
                    log(f"       Tab '{sname}': A1 date update only")

            wb_w.save(dest)

        else:  # .xlsx
            shutil.copy2(src, dest)
            wb = openpyxl.load_workbook(dest)

            for sname in wb.sheetnames:
                ws = wb[sname]

                # FIX GAP 2
                _supp_a1_update_xlsx(ws, new_date_str, sname)

                if sname == "Total Loan":
                    log(f"       Processing Total Loan tab...")
                    _process_total_loan_xlsx(ws, loan_map, irp_data)
                elif sname == "Comp Finan Status":
                    log(f"       Processing Comp Finan Status tab...")
                    _process_comp_finan_xlsx(ws, loan_map, irp_data)
                elif sname == "Res LOC Report":
                    log(f"       Processing Res LOC Report tab...")
                    if pirp_data:
                        _process_res_loc_xlsx(ws, pirp_data, trans_id)
                    else:
                        log("       Res LOC: No PIRPXLLR data - carried forward as-is")
                elif sname == "Watchlist":
                    # Watchlist: A1 date update only.
                    # Watchlist data is managed by the servicer — do not inject formulas.
                    log(f"       Tab 'Watchlist': A1 date update only (no formula injection)")

            wb.save(dest)

        log(f"     Supplemental: SAVED -> {dest}")
        for w in validate_output_file(dest, "Supplemental"):
            log(f"     {w}", "WARN")
        return dest, "", {}
    except Exception as e:
        return "", str(e), {}

# ── Financial file ────────────────────────────────────────────────────────────

def create_financial(trans_id, det_date, prev_folder, out_folder, excel_app=None):
    """
    Copy-forward Financial file, update A1 dates on all tabs.
    FIX GAP 5: Uses replace_date_in_text (not raw overwrite).
               Windows-safe date format (%m/%d/%Y, 4-digit year).
               Handles both pure date cells and text cells.
    """
    src = find_financial_file(prev_folder, trans_id)

    if not src and cfg.TEST_MODE:
        prod_prev = prod_path(prev_folder)
        if prod_prev != prev_folder and os.path.isdir(prod_prev):
            src = find_financial_file(prod_prev, trans_id)
            if src:
                log(f"     Financial: Production fallback -> {os.path.basename(src)}")

    if not src:
        log(f"     Financial: SKIP — no Financial file found in {prev_folder}")
        return "", "No Financial file found", {}

    clean = clean_filename(trans_id)
    dt    = parse_det_date(det_date)
    if dt:
        prefix       = dt.strftime("%Y.%m")
        new_date_str = format_date_long(dt)   # mm/dd/yyyy — matches VBA Financial format
    else:
        prefix       = det_date[:7]
        new_date_str = det_date

    src_ext = os.path.splitext(src)[1].lower()
    if src_ext == ".xls" and excel_app is not None:
        converted_src = os.path.join(out_folder, f"{prefix} CREFC Financial File {clean}_source.xlsx")
        xls_to_xlsx(src, converted_src, excel_app)
        src = converted_src
        src_ext = ".xlsx"
    dest    = os.path.join(out_folder, f"{prefix} CREFC Financial File {clean}.xlsx")
    log(f"     Financial: Source file: {os.path.basename(src)}")

    try:
        if src_ext == ".xls":
            rb   = xlrd.open_workbook(src, formatting_info=True)
            wb_w = xl_copy(rb)

            for sheet_idx in range(rb.nsheets):
                ws_r  = rb.sheet_by_index(sheet_idx)
                ws_w  = wb_w.get_sheet(sheet_idx)
                sname = ws_r.name
                a1    = ws_r.cell(0, 0)

                if a1.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                elif a1.ctype == xlrd.XL_CELL_DATE:
                    # FIX GAP 5: Pure date cell — write the date object directly
                    ws_w.write(0, 0, dt if dt else new_date_str)
                    log(f"       Tab '{sname}': A1 date updated to {new_date_str}")
                else:
                    # Text cell — replace date in text, preserve surrounding content
                    a1_text = str(a1.value or "")
                    updated = replace_date_in_text(a1_text, new_date_str)
                    if updated != a1_text:
                        ws_w.write(0, 0, updated)
                        log(f"       Tab '{sname}': A1 text updated to {new_date_str}")

            wb_w.save(dest)

        else:  # .xlsx
            shutil.copy2(src, dest)
            wb = openpyxl.load_workbook(dest)

            for sname in wb.sheetnames:
                ws   = wb[sname]
                a1   = ws["A1"]
                if a1.value is None:
                    continue
                if isinstance(a1.value, (datetime, date)):
                    # Pure date cell
                    ws["A1"] = dt if dt else new_date_str
                    _highlight_cell(ws["A1"])
                    log(f"       Tab '{sname}': A1 date updated to {new_date_str}")
                else:
                    a1_text = str(a1.value)
                    updated = replace_date_in_text(a1_text, new_date_str)
                    if updated != a1_text:
                        ws["A1"] = updated
                        _highlight_cell(ws["A1"])
                        log(f"       Tab '{sname}': A1 text updated to {new_date_str}")

            wb.save(dest)

        log(f"     Financial: SAVED -> {dest}")
        for w in validate_output_file(dest, "Financial"):
            log(f"     {w}", "WARN")
        return dest, "", {}
    except Exception as e:
        return "", str(e), {}

# ── Excel run log writer ──────────────────────────────────────────────────────

STATUS_COLORS = {
    "CREATED": "C6EFCE",   # green
    "UPDATED": "FFEB9C",   # yellow
    "SKIPPED": "FFCCCC",   # light red
    "ERROR":   "FF0000",   # red
}

def write_excel_log(log_path, deal_summary, elapsed_sec, filtered_deals=None):
    """Write 4-tab Excel log: Deal Summary, File Detail, Filtered Deals, Run Log."""
    if filtered_deals is None:
        filtered_deals = []
    wb = openpyxl.Workbook()

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    thin     = Side(style="thin")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_fill = PatternFill("solid", fgColor="FFCCCC")

    # ── Tab 1: Deal Summary ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Deal Summary"

    cols1   = ["Deal", "Det Date", "Servicer", "Status", "Files Created",
               "Periodic Matched", "Periodic Unmatched", "Property Matched", "Property Unmatched",
               "Periodic Path", "Property Path", "Supplemental Path", "Financial Path",
               "Notes / Error"]
    widths1 = [22, 12, 10, 10, 14, 16, 18, 16, 18, 60, 60, 60, 60, 50]

    for c, h in enumerate(cols1, 1):
        cell = ws1.cell(2, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for c, w in enumerate(widths1, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    for i, deal in enumerate(deal_summary, 3):
        p_metrics = deal.get("periodic_metrics", {})
        pr_metrics = deal.get("property_metrics", {})
        row_data = [
            deal["trans_id"], deal["det_date"], deal["servicer"], deal["status"],
            deal["files_count"],
            p_metrics.get("matched", ""), p_metrics.get("unmatched", ""),
            pr_metrics.get("matched", ""), pr_metrics.get("unmatched", ""),
            deal.get("periodic_path", ""), deal.get("property_path", ""),
            deal.get("supplemental_path", ""), deal.get("financial_path", ""),
            deal.get("note", ""),
        ]
        color = STATUS_COLORS.get(deal["status"].upper(), "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(i, c, val)
            cell.border = border
            if c == 4:
                cell.fill = fill
            # Highlight unmatched columns red if > 0
            if c in (7, 9) and isinstance(val, int) and val > 0:
                cell.fill = red_fill
            cell.alignment = Alignment(wrap_text=(c >= 10))

    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(cols1))}2"

    # Summary banner in row 1
    created     = sum(1 for d in deal_summary if d["status"].upper() == "CREATED")
    skipped     = sum(1 for d in deal_summary if d["status"].upper() == "SKIPPED")
    errors      = sum(1 for d in deal_summary if d["status"].upper() == "ERROR")
    total_files = sum(d["files_count"] for d in deal_summary)
    total_unmatched = sum(
        d.get(f"{ft}_metrics", {}).get("unmatched", 0)
        for d in deal_summary for ft in ["periodic", "property"]
    )
    banner = (f"CMBS Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
              f"Deals: {len(deal_summary)}  |  Created: {created}  |  "
              f"Skipped: {skipped}  |  Errors: {errors}  |  "
              f"Files written: {total_files}  |  "
              f"Filtered: {len(filtered_deals)}  |  "
              f"Unmatched: {total_unmatched}  |  Elapsed: {elapsed_sec:.0f}s")
    banner_cell = ws1.cell(1, 1, banner)
    banner_cell.font = Font(bold=True, size=11)
    if total_unmatched > 0 or errors > 0:
        banner_cell.font = Font(bold=True, size=11, color="FF0000")
    ws1.merge_cells(f"A1:{openpyxl.utils.get_column_letter(len(cols1))}1")
    ws1.row_dimensions[1].height = 20

    # ── Tab 2: File Detail ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("File Detail")
    cols2   = ["Deal", "Det Date", "Servicer", "File Type", "Status",
               "Matched", "Unmatched", "Total Rows", "File Path", "Notes"]
    widths2 = [22, 12, 10, 16, 10, 10, 12, 12, 90, 50]

    for c, h in enumerate(cols2, 1):
        cell = ws2.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for c, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    for i, fr in enumerate(file_rows, 2):
        row_data = [fr["deal"], fr["det_date"], fr["servicer"],
                    fr["file_type"], fr["status"],
                    fr.get("matched", ""), fr.get("unmatched", ""),
                    fr.get("total_rows", ""),
                    fr["path"], fr["note"]]
        color = STATUS_COLORS.get(fr["status"].upper(), "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(i, c, val)
            cell.border = border
            if c == 5:
                cell.fill = fill
            if c == 7 and isinstance(val, int) and val > 0:
                cell.fill = red_fill
            cell.alignment = Alignment(wrap_text=(c == 9))

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols2))}1"

    # ── Tab 3: Filtered Deals ────────────────────────────────────────────────
    ws_f = wb.create_sheet("Filtered Deals")
    cols_f = ["Deal", "Filter Reason"]
    for c, h in enumerate(cols_f, 1):
        cell = ws_f.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
    ws_f.column_dimensions["A"].width = 30
    ws_f.column_dimensions["B"].width = 40

    for i, (tid, reason) in enumerate(filtered_deals, 2):
        ws_f.cell(i, 1, tid).border = border
        ws_f.cell(i, 2, reason).border = border

    if not filtered_deals:
        ws_f.cell(2, 1, "(none)").font = Font(italic=True)

    ws_f.freeze_panes = "A2"

    # ── Tab 4: Run Log ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Run Log")
    for c, h in enumerate(["Time", "Level", "Message"], 1):
        cell = ws3.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 130

    for i, entry in enumerate(log_rows, 2):
        ws3.cell(i, 1, entry["time"])
        ws3.cell(i, 2, entry["level"])
        ws3.cell(i, 3, entry["msg"])
        if entry["level"] == "ERROR":
            for c in range(1, 4):
                ws3.cell(i, c).fill = PatternFill("solid", fgColor="FF0000")
                ws3.cell(i, c).font = Font(color="FFFFFF")

    ensure_dir(os.path.dirname(log_path))
    wb.save(log_path)
    log(f"  Log saved: {log_path}")

# ── Main ──────────────────────────────────────────────────────────────────────

def find_irp_folder():
    print("\n" + "=" * 70)
    print("  CMBS INVESTOR REPORTING TOOL  (Python Edition v2.4)")
    if cfg.TEST_MODE:
        print(f"  *** TEST MODE \u2014 Output: {cfg.TEST_ROOT} ***")
    print("=" * 70)

    # Auto-detect: find the most recent "Investor Reporting Package" folder
    # in the user's Downloads directory
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    candidates = []
    if os.path.isdir(downloads):
        for name in os.listdir(downloads):
            if "Investor Reporting Package" in name:
                full = os.path.join(downloads, name)
                if os.path.isdir(full):
                    candidates.append((os.path.getmtime(full), full, name))

    if candidates:
        candidates.sort(reverse=True)   # newest first
        auto_folder = candidates[0][1]
        auto_name   = candidates[0][2]
        print(f"\n  Auto-detected IRP folder:")
        print(f"    {auto_name}")
        if len(candidates) > 1:
            print(f"    ({len(candidates)} matching folders found - using most recent)")
        print("\n  Press Enter to use this, or paste a different path to override:")
        answer = input("  > ").strip().strip('"').strip("'")
        if not answer:
            return auto_folder
        return answer

    # Fallback: no matching folder found
    print("\n  No 'Investor Reporting Package' folder found in Downloads.")
    print("  Paste the IRP folder path (or press Enter to browse):")
    folder = input("  > ").strip().strip('"').strip("'")
    if not folder:
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk(); root.withdraw()
            folder = filedialog.askdirectory(title="Select IRP Download Folder")
        except Exception:
            print("ERROR: No folder selected.")
            sys.exit(1)
    return folder


def run():
    global run_start
    run_start = datetime.now()

    parser = argparse.ArgumentParser(description="CMBS Investor Reporting Tool")
    parser.add_argument("--dry-run", action="store_true",
                        help="Resolve all deal paths and report without writing files")
    args, _ = parser.parse_known_args()
    dry_run = args.dry_run

    folder = find_irp_folder()
    if not os.path.isdir(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    log(f"  Selected folder: {folder}")

    # Find GANDATA / PIRPXLPU and PIRPXLLR files
    gandata_files = (glob.glob(os.path.join(folder, "*GANDATA*")) +
                     glob.glob(os.path.join(folder, "*PIRPXLPU*")))
    pirp_files    =  glob.glob(os.path.join(folder, "*PIRPXLLR*"))

    if not gandata_files:
        print("ERROR: No GANDATA/PIRPXLPU file found in folder.")
        sys.exit(1)

    irp_path  = gandata_files[0]
    pirp_path = pirp_files[0] if pirp_files else None

    log(f"  GANDATA (IRP): {os.path.basename(irp_path)}")
    if pirp_path:
        log(f"  PIRPXLLR:      {os.path.basename(pirp_path)}")
    else:
        log("  No PIRPXLLR file found - Res LOC will carry forward")

    irp_data  = read_irp(irp_path)
    validate_irp_columns(irp_data)
    validate_overrides()
    sync_tracking_list()
    pirp_data = read_pirpxllr(pirp_path) if pirp_path else []

    # Start a shared Excel COM instance for .xls -> .xlsx conversion, when available.
    excel_app = None
    if win32 is not None and sys.platform.startswith("win"):
        try:
            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            log("  Excel COM automation initialised for .xls conversion")
        except Exception as e:
            log_err(f"  Excel COM automation could not be started: {e}")
            excel_app = None

    trans_map = build_trans_map(irp_data)
    log(f"\n  Unique deals before filter: {len(trans_map)}")

    # Step 1: remove zero-balance (paid off) deals
    trans_map, zero_bal_removed = filter_zero_balance(trans_map, irp_data)

    # Step 2: remove deals with no Paid Through Date in PIRPXLLR
    if pirp_data:
        active_map, ptd_removed = filter_by_pirpxllr(trans_map, pirp_data)
    else:
        active_map = trans_map
        ptd_removed = []

    all_filtered = zero_bal_removed + ptd_removed

    log(f"  Active deals after filter:  {len(active_map)}\n")

    sorted_deals = sorted(active_map.keys())
    print("\nDeals to process:")
    for i, tid in enumerate(sorted_deals, 1):
        svc = get_servicer(irp_data, tid)
        print(f"  {i:2d}. {tid}  [{svc}]")

    # ── Dry-run mode: resolve paths and report without writing files ──
    if dry_run:
        print("\n" + "=" * 60)
        print("  DRY RUN: Path resolution check (no files will be written)")
        print("=" * 60)
        issues = 0
        for tid in sorted_deals:
            servicer = get_servicer(irp_data, tid)
            det_date = get_det_date(irp_data, tid)
            lender = servicer_to_folder(servicer)
            output = resolve_output_folder(tid, servicer)
            crefc_parent = ""
            if output:
                try:
                    out_folder = build_crefc_folder(output, det_date)
                    crefc_parent = os.path.dirname(out_folder.rstrip(os.sep)) + os.sep
                except Exception:
                    pass
            prev = find_prev_month_folder(crefc_parent, det_date) if crefc_parent else ""
            ok = bool(output and prev)
            if not ok:
                issues += 1
            print(f"\n  {'OK' if ok else 'ISSUE':>5}  {tid}")
            print(f"         Servicer: {servicer} -> {lender}")
            print(f"         Output:   {output or 'NOT FOUND'}")
            print(f"         Previous: {prev or 'NOT FOUND'}")
            if not output:
                print(f"         FIX: Add to deal_overrides.json folder_overrides")
            elif not prev:
                print(f"         FIX: Ensure previous month CREFC folder exists in {crefc_parent}")

        if all_filtered:
            print(f"\n  Filtered out ({len(all_filtered)} deals):")
            for tid, reason in all_filtered:
                print(f"    {tid}: {reason}")

        print(f"\n  Summary: {len(sorted_deals)} active deals, {issues} with path issues, {len(all_filtered)} filtered")
        sys.exit(0)

    print(f"\nProcess all {len(sorted_deals)} deals? (Y/n): ", end="")
    if input().strip().lower() == "n":
        print("Aborted.")
        sys.exit(0)

    deal_summary  = []
    files_created = 0

    for idx, trans_id in enumerate(sorted_deals, 1):
        det_date = get_det_date(irp_data, trans_id)
        servicer = get_servicer(irp_data, trans_id)

        print(f"\n  [{idx}/{len(sorted_deals)}] {trans_id}")
        log(f"\n  -- Processing: {trans_id} --")
        log(f"     Det Date: {det_date}   Servicer: {servicer}")

        deal_record = {
            "trans_id": trans_id, "det_date": det_date, "servicer": servicer,
            "status": "CREATED", "files_count": 0, "note": "",
        }

        output_folder = resolve_output_folder(trans_id, servicer)
        if not output_folder:
            reason = f"Could not resolve S: drive folder (servicer='{servicer}')"
            log_err(reason)
            deal_record.update({"status": "SKIPPED", "note": reason})
            add_file_row(trans_id, det_date, servicer, "ALL", "SKIPPED", "", reason)
            deal_summary.append(deal_record)
            continue

        try:
            out_folder = build_crefc_folder(output_folder, det_date)
        except Exception as e:
            deal_record.update({"status": "ERROR", "note": str(e)})
            deal_summary.append(deal_record)
            continue

        log(f"     Output: {out_folder}")

        crefc_parent = os.path.dirname(out_folder.rstrip(os.sep)) + os.sep
        prev_folder  = find_prev_month_folder(crefc_parent, det_date)
        if prev_folder:
            log(f"     Previous month: {prev_folder}")
        else:
            log("     No previous month folder found - all 4 files will be skipped")

        deal_files = 0

        for file_type, fn in [
            ("Periodic",     lambda: create_periodic(trans_id, det_date, irp_data, pirp_data, prev_folder, out_folder, excel_app)),
            ("Property",     lambda: create_property(trans_id, det_date, irp_data, prev_folder, out_folder, excel_app)),
            ("Supplemental", lambda: create_supplemental(trans_id, det_date, irp_data, pirp_data, prev_folder, out_folder, excel_app)),
            ("Financial",    lambda: create_financial(trans_id, det_date, prev_folder, out_folder, excel_app)),
        ]:
            if prev_folder:
                path, err, metrics = fn()
            else:
                path, err, metrics = "", "No previous month folder", {}
            status = "CREATED" if path else "SKIPPED"
            add_file_row(trans_id, det_date, servicer, file_type, status, path, err,
                         matched=metrics.get("matched", 0),
                         unmatched=metrics.get("unmatched", 0),
                         total_rows=metrics.get("total_rows", 0))
            deal_record[f"{file_type.lower()}_path"] = path
            deal_record[f"{file_type.lower()}_metrics"] = metrics
            if path:
                deal_files += 1

        deal_record["files_count"] = deal_files
        files_created += deal_files
        if deal_files == 0:
            deal_record["status"] = "SKIPPED"
            deal_record["note"]   = "No prior month files to copy forward"
        deal_summary.append(deal_record)

    # Integrity check
    if len(deal_summary) != len(sorted_deals):
        log_err(f"  INTEGRITY: Expected {len(sorted_deals)} deals, only {len(deal_summary)} in summary")

    total_unmatched = sum(
        d.get(f"{ft}_metrics", {}).get("unmatched", 0)
        for d in deal_summary for ft in ["periodic", "property"]
    )
    if total_unmatched > 0:
        log(f"  WARNING: {total_unmatched} total unmatched loan rows across all deals", "WARN")

    # Write log
    elapsed  = (datetime.now() - run_start).total_seconds()
    ts       = run_start.strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(cfg.LOG_FOLDER, f"CMBS_Log_{ts}.xlsx")

    log(f"\n{'='*60}")
    log(f"  RUN COMPLETE")
    log(f"  Deals processed: {len(deal_summary)}")
    log(f"  Files written:   {files_created}")
    log(f"  Filtered:        {len(all_filtered)}")
    if total_unmatched > 0:
        log(f"  Unmatched loans: {total_unmatched}")
    log(f"  Elapsed:         {elapsed:.1f}s")

    write_excel_log(log_path, deal_summary, elapsed, all_filtered)

    # Clean up Excel COM instance if used
    if excel_app is not None:
        try:
            excel_app.Quit()
        except Exception:
            pass

    print(f"\n  Opening log: {log_path}")
    os.startfile(log_path)


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\nCancelled.")
    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        traceback.print_exc()
    input("\nPress Enter to close...")
